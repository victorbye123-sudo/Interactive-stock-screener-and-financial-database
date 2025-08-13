import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Literal, Any, Optional, Dict
from contextlib import asynccontextmanager
from functools import lru_cache
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import StringIO
import io
import csv
import re

from pathlib import Path
import os

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = Path(os.getenv('DUCKDB_PATH', BASE_DIR / 'screener.duckdb'))

import duckdb
import pandas as pd
import requests
import yfinance as yf
from fastapi import FastAPI, HTTPException, Body, Query, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
from starlette.middleware.base import BaseHTTPMiddleware
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# --------- Fast JSON (fallback to std if orjson unavailable) ----------
try:
    from fastapi.responses import ORJSONResponse as DefaultJSONResponse
except Exception:  # pragma: no cover
    from fastapi.responses import JSONResponse as DefaultJSONResponse

# ---------- HTTP session ----------
HEADERS = {
    "User-Agent": "Victor Bye victorbye123@gmail.com (for research; contact if issues)",
    "Accept-Encoding": "gzip, deflate",
}
SESSION = requests.Session()
SESSION.headers.update(HEADERS)
_adapter = HTTPAdapter(
    pool_connections=100,
    pool_maxsize=100,
    max_retries=Retry(total=3, backoff_factor=0.3, status_forcelist=[429, 500, 502, 503, 504]),
)
SESSION.mount("https://", _adapter)
SESSION.mount("http://", _adapter)
TIMEOUT = 20

# ---------- Dev-friendly cache control ----------
NO_CACHE = os.getenv("NO_CACHE", "1") == "1"
NO_CACHE_HEADERS = (
    {"Cache-Control": "no-store, max-age=0, must-revalidate"}
    if NO_CACHE
    else {"Cache-Control": "public, max-age=60"}
)


def _get_json(url: str):
    r = SESSION.get(url, timeout=TIMEOUT)
    r.raise_for_status()
    return r.json()


def _get_text(url: str) -> str:
    r = SESSION.get(url, timeout=TIMEOUT)
    r.raise_for_status()
    r.encoding = "utf-8"
    return r.text


# ---------- Security headers ----------
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        resp = await call_next(request)
        resp.headers.setdefault("X-Content-Type-Options", "nosniff")
        resp.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
        resp.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
        resp.headers.setdefault("Permissions-Policy", "geolocation=(), microphone=(), camera=()")
        # HSTS only if served over HTTPS behind a proxy/terminator
        if request.url.scheme == "https":
            resp.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
        return resp


# ---------- App & DB lifespan ----------
def _init_db(con: duckdb.DuckDBPyConnection) -> None:
    # Create tables if they don't exist
    con.execute("""
                CREATE TABLE IF NOT EXISTS metrics
                (
                    ticker
                    TEXT
                    PRIMARY
                    KEY,
                    name
                    TEXT,
                    sector
                    TEXT,
                    industry
                    TEXT,
                    price
                    DOUBLE,
                    shares
                    BIGINT,
                    market_cap
                    DOUBLE,
                    pe
                    DOUBLE,
                    ps
                    DOUBLE,
                    pb
                    DOUBLE,
                    revenue_fy
                    DOUBLE,
                    net_income_fy
                    DOUBLE,
                    equity_fy
                    DOUBLE,
                    debt
                    DOUBLE,
                    debt_to_equity
                    DOUBLE,
                    rev_cagr_3y
                    DOUBLE,
                    net_margin
                    DOUBLE,
                    updated_at
                    TIMESTAMP
                );
                """)
    con.execute("""
                CREATE TABLE IF NOT EXISTS symbols
                (
                    ticker
                    TEXT
                    PRIMARY
                    KEY,
                    name
                    TEXT,
                    exchange
                    TEXT,
                    country
                    TEXT,
                    currency
                    TEXT,
                    sector
                    TEXT,
                    industry
                    TEXT,
                    active
                    BOOLEAN
                );
                """)
    # 🔧 Ensure required columns exist for both tables (adds any that are missing)
    _ensure_columns(con, "symbols", {
        "name": "TEXT", "exchange": "TEXT", "country": "TEXT", "currency": "TEXT",
        "sector": "TEXT", "industry": "TEXT", "active": "BOOLEAN"
    })
    _ensure_columns(con, "metrics", {
        "name": "TEXT", "sector": "TEXT", "industry": "TEXT", "price": "DOUBLE",
        "shares": "BIGINT", "market_cap": "DOUBLE", "pe": "DOUBLE", "ps": "DOUBLE",
        "pb": "DOUBLE", "revenue_fy": "DOUBLE", "net_income_fy": "DOUBLE",
        "equity_fy": "DOUBLE", "debt": "DOUBLE", "debt_to_equity": "DOUBLE",
        "rev_cagr_3y": "DOUBLE", "net_margin": "DOUBLE", "updated_at": "TIMESTAMP"
    })
    # ✅ Build indexes only after columns are guaranteed to exist
    _ensure_indexes(con)


@asynccontextmanager
async def lifespan(app: FastAPI):
    con = duckdb.connect(str(DB_PATH))
    _init_db(con)
    try:
        con.execute("PRAGMA threads=4")
    except Exception:
        pass
    app.state.con = con
    try:
        yield
    finally:
        try:
            con.close()
        except Exception:
            pass


app = FastAPI(
    title="Global Screener + Financials",
    lifespan=lifespan,
    default_response_class=DefaultJSONResponse,
)
app.add_middleware(GZipMiddleware, minimum_size=1024)
app.add_middleware(SecurityHeadersMiddleware)

# Optional CORS
if os.getenv("ALLOW_CORS", "0") == "1":
    origins = [o.strip() for o in os.getenv("CORS_ORIGINS", "*").split(",") if o.strip()]
    app.add_middleware(
        CORSMiddleware,
        allow_origins=origins,
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )


# ---------- DuckDB helpers ----------
def _table_cols(con: duckdb.DuckDBPyConnection, table: str) -> List[str]:
    # PRAGMA table_info returns: [cid, name, type, notnull, dflt_value, pk]
    return [r[1] for r in con.execute(f"PRAGMA table_info('{table}')").fetchall()]


def _upsert_df(con: duckdb.DuckDBPyConnection, table: str, df: pd.DataFrame) -> int:
    """Generic INSERT OR REPLACE using DuckDB; respects table column order."""
    if df is None or df.empty:
        return 0
    cols = _table_cols(con, table)
    use_cols = [c for c in cols if c in df.columns]
    tmp = f"tmp_{table}"
    con.register(tmp, df[use_cols])
    con.execute(
        f"INSERT OR REPLACE INTO {table} ({','.join(use_cols)}) SELECT {','.join(use_cols)} FROM {tmp}"
    )
    con.unregister(tmp)
    return int(len(df))


def _ensure_columns(con: duckdb.DuckDBPyConnection, table: str, required: Dict[str, str]) -> None:
    """Add any missing columns to an existing DuckDB table."""
    cols = {r[1]: r[2] for r in con.execute(f"PRAGMA table_info('{table}')").fetchall()}  # name -> type
    for col, typ in required.items():
        if col not in cols:
            con.execute(f'ALTER TABLE {table} ADD COLUMN "{col}" {typ}')


def _ensure_indexes(con: duckdb.DuckDBPyConnection) -> None:
    """Create indexes only if the target columns exist."""
    have_metrics = {r[1] for r in con.execute("PRAGMA table_info('metrics')").fetchall()}
    if "sector" in have_metrics:
        con.execute("CREATE INDEX IF NOT EXISTS idx_metrics_sector ON metrics(sector)")
    if "updated_at" in have_metrics:
        con.execute("CREATE INDEX IF NOT EXISTS idx_metrics_updated ON metrics(updated_at)")

    have_symbols = {r[1] for r in con.execute("PRAGMA table_info('symbols')").fetchall()}
    if "sector" in have_symbols:
        con.execute("CREATE INDEX IF NOT EXISTS idx_symbols_sector ON symbols(sector)")
    if "active" in have_symbols:
        con.execute("CREATE INDEX IF NOT EXISTS idx_symbols_active ON symbols(active)")


# ---------- SEC helpers (US only) ----------
@lru_cache(maxsize=1)
def _company_map() -> Dict[str, tuple]:
    try:
        data = _get_json("https://www.sec.gov/files/company_tickers.json")
        return {v["ticker"].upper(): (str(v["cik_str"]).zfill(10), v["title"]) for v in data.values()}
    except Exception as e:
        print("SEC company map load failed:", e)
        return {}


def _fy_series(facts: dict, tag: str) -> pd.Series:
    obj = facts.get("facts", {}).get("us-gaap", {}).get(tag, {})
    for unit in ("USD", "USD/shares", "shares"):
        arr = obj.get("units", {}).get(unit)
        if not arr:
            continue
        vals = {}
        for row in arr:
            if row.get("fp") != "FY":
                continue
            v = row.get("val")
            if isinstance(v, (int, float)):
                d = row.get("end") or row.get("instant")
                if d and len(d) >= 4:
                    vals[d[:4]] = float(v)
        if vals:
            return pd.Series(vals, dtype="float64").sort_index()
    return pd.Series(dtype="float64")


def _try_first(facts: dict, tags: List[str]) -> pd.Series:
    for t in tags:
        s = _fy_series(facts, t)
        if not s.empty:
            return s
    return pd.Series(dtype="float64")


@lru_cache(maxsize=4096)
def _companyfacts(cik: str):
    return _get_json(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json")


# ---------- Symbol resolution / normalization ----------
def _normalize_ticker(sym: str) -> str:
    s = (sym or "").strip().upper()
    if "." in s and (s.endswith(".A") or s.endswith(".B") or s.endswith(".C")) and "-" not in s and not s.endswith(
            ".L") and not s.endswith(".TO"):
        parts = s.split(".")
        if len(parts) == 2 and len(parts[1]) == 1:
            return parts[0] + "-" + parts[1]
    return s


def _sec_lookup_key(sym: str) -> str:
    s = (sym or "").strip().upper()
    if "-" in s:
        base, *rest = s.split("-")
        if len(rest) == 1 and len(rest[0]) == 1 and base.isalpha():
            return f"{base}.{rest[0]}"
    return s


def _yahoo_suggest(q: str, quotes_count: int = 10) -> List[Dict[str, Any]]:
    url = "https://query2.finance.yahoo.com/v1/finance/search"
    try:
        r = SESSION.get(url, params={"q": q, "quotesCount": quotes_count, "newsCount": 0, "listsCount": 0},
                        timeout=TIMEOUT)
        r.raise_for_status()
        data = r.json() or {}
    except Exception:
        return []
    out = []
    for it in data.get("quotes", []) or []:
        sym = it.get("symbol")
        if not sym:
            continue
        out.append(
            {
                "symbol": sym,
                "shortname": it.get("shortname") or "",
                "longname": it.get("longname") or "",
                "exchange": it.get("exchDisp") or "",
                "sector": it.get("sector") or "",
                "industry": it.get("industry") or "",
                "quoteType": it.get("quoteType") or "",
            }
        )
    return out


@app.get("/resolve")
def resolve(q: str = Query(..., description="Ticker or company name")):
    qn = _normalize_ticker(q)
    items = _yahoo_suggest(qn)
    return {"query": q, "normalized": qn, "matches": items[:20]}


@app.get("/suggest")
def suggest(q: str = Query(..., description="Ticker or company name"), limit: int = 10):
    items = _yahoo_suggest(q, quotes_count=max(1, min(int(limit), 25)))
    return {
        "matches": [
            {
                "symbol": it.get("symbol"),
                "name": it.get("longname") or it.get("shortname") or "",
                "exchange": it.get("exchange") or "",
                "sector": it.get("sector") or "",
                "industry": it.get("industry") or "",
            }
            for it in (items or [])[:limit]
        ]
    }


# ---------- Yahoo helpers (global) ----------
def _price_fast(yf_tkr: yf.Ticker) -> Optional[float]:
    info = getattr(yf_tkr, "fast_info", {}) or {}
    p = info.get("last_price")
    if p:
        try:
            return float(p)
        except Exception:
            pass
    try:
        # last known close if fast_info missing (ETF/foreign)
        return float(yf_tkr.history(period="1d")["Close"].iloc[-1])
    except Exception:
        return None


@lru_cache(maxsize=4096)
def _yahoo_basic(ticker: str) -> dict:
    t = yf.Ticker(ticker)
    fi = getattr(t, "fast_info", {}) or {}
    out = {
        "price": _price_fast(t),
        "market_cap": fi.get("market_cap"),
        "shares": fi.get("shares"),
        "sector": None,
        "industry": None,
        "name": None,
        "pe": None,
        "ps": None,
        "pb": None,
        "revenue_fy": None,
        "net_income_fy": None,
        "equity_fy": None,
        "debt": None,
        "longTermDebt": None,
        "shortLongTermDebt": None,
    }
    info = {}
    try:
        # yfinance.get_info can be slow/flaky; only call once and swallow failures
        info = t.get_info() or {}
    except Exception:
        info = {}
    out.update(
        {
            "sector": info.get("sector"),
            "industry": info.get("industry"),
            "name": info.get("longName") or info.get("shortName"),
            "shares": out["shares"] or info.get("sharesOutstanding"),
            "market_cap": out["market_cap"] or info.get("marketCap"),
            "pe": info.get("trailingPE") or info.get("forwardPE"),
            "ps": info.get("priceToSalesTrailing12Months"),
            "pb": info.get("priceToBook"),
            "revenue_fy": info.get("totalRevenue"),
            "net_income_fy": info.get("netIncomeToCommon") or info.get("netIncome"),
            "equity_fy": info.get("totalStockholderEquity"),
            "debt": info.get("totalDebt"),
            "longTermDebt": info.get("longTermDebt"),
            "shortLongTermDebt": info.get("shortLongTermDebt"),
        }
    )
    return out


def _metrics_from_yahoo_only(ticker: str) -> dict:
    ticker = _normalize_ticker(ticker)
    y = _yahoo_basic(ticker)
    price = y["price"]
    shares = y["shares"]
    market_cap = y["market_cap"] or (price * shares if price and shares else None)
    debt = y["debt"]
    if debt is None:
        lt = y.get("longTermDebt") or 0
        st = y.get("shortLongTermDebt") or 0
        try:
            debt = float(lt) + float(st)
        except Exception:
            debt = None
    dte = (debt / y["equity_fy"]) if (debt not in (None, 0) and y["equity_fy"] not in (None, 0)) else None
    net_margin = (
        (y["net_income_fy"] / y["revenue_fy"])
        if (y["net_income_fy"] not in (None, 0) and y["revenue_fy"] not in (None, 0))
        else None
    )
    rev_cagr_3y = None
    try:
        t = yf.Ticker(ticker)
        try:
            is_df = t.get_income_stmt(freq="yearly")
        except Exception:
            is_df = t.get_income_stmt(freq="annual")
        if isinstance(is_df, pd.DataFrame) and "Total Revenue" in is_df.index:
            s = is_df.loc["Total Revenue"].dropna().sort_index()
            if len(s) >= 4:
                start, end = float(s.iloc[-4]), float(s.iloc[-1])
                if start > 0:
                    rev_cagr_3y = (end / start) ** (1 / 3) - 1
    except Exception:
        pass
    return {
        "ticker": ticker,
        "name": y["name"] or ticker.upper(),
        "sector": y["sector"],
        "industry": y["industry"],
        "price": float(price) if price is not None else None,
        "shares": int(shares) if shares else None,
        "market_cap": float(market_cap) if market_cap is not None else None,
        "pe": float(y["pe"]) if y["pe"] is not None else None,
        "ps": float(y["ps"]) if y["ps"] is not None else None,
        "pb": float(y["pb"]) if y["pb"] is not None else None,
        "revenue_fy": float(y["revenue_fy"]) if y["revenue_fy"] is not None else None,
        "net_income_fy": float(y["net_income_fy"]) if y["net_income_fy"] is not None else None,
        "equity_fy": float(y["equity_fy"]) if y["equity_fy"] is not None else None,
        "debt": float(debt) if debt is not None else None,
        "debt_to_equity": float(dte) if dte is not None else None,
        "rev_cagr_3y": float(rev_cagr_3y) if rev_cagr_3y is not None else None,
        "net_margin": float(net_margin) if net_margin is not None else None,
        "updated_at": datetime.utcnow(),
    }


# ---------- Metric computation ----------
def compute_metrics_for(ticker: str) -> dict:
    yahoo_sym = _normalize_ticker(ticker)
    sec_key = _sec_lookup_key(ticker)
    cmap = _company_map()
    if sec_key not in cmap:
        return _metrics_from_yahoo_only(yahoo_sym)
    cik, legal_name = cmap[sec_key]
    facts = _companyfacts(cik)
    rev = _try_first(
        facts,
        [
            "RevenueFromContractWithCustomerExcludingAssessedTax",
            "Revenues",
            "SalesRevenueNet",
            "TotalRevenue",
        ],
    )
    ni = _try_first(facts, ["NetIncomeLoss", "ProfitLoss"])
    eq = _try_first(
        facts,
        [
            "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest",
            "StockholdersEquity",
            "Equity",
        ],
    )
    debt_cur = _try_first(facts, ["DebtCurrent", "ShortTermBorrowings", "CommercialPaper"])
    debt_lt = _try_first(facts, ["LongTermDebtNoncurrent", "LongTermDebt"])
    debt = (debt_cur.add(debt_lt, fill_value=0)).sort_index()

    rev_cagr_3y = None
    if len(rev) >= 4:
        start, end = rev.iloc[-4], rev.iloc[-1]
        if start and start > 0:
            rev_cagr_3y = (end / start) ** (1 / 3) - 1
    revenue_fy = float(rev.iloc[-1]) if not rev.empty else None
    net_income_fy = float(ni.iloc[-1]) if not ni.empty else None
    equity_fy = float(eq.iloc[-1]) if not eq.empty else None
    debt_latest = float(debt.iloc[-1]) if not debt.empty else None
    net_margin = (net_income_fy / revenue_fy) if (net_income_fy is not None and revenue_fy not in (None, 0)) else None

    y = _yahoo_basic(yahoo_sym)
    price = y["price"]
    sector = y["sector"]
    industry = y["industry"]
    shares = y["shares"]
    market_cap = (price * shares) if (price and shares) else y["market_cap"]
    eps = (net_income_fy / shares) if (net_income_fy and shares) else None
    pe = (price / eps) if (price and eps not in (None, 0)) else y["pe"]
    ps = (market_cap / revenue_fy) if (market_cap and revenue_fy not in (None, 0)) else y["ps"]
    pb = (market_cap / equity_fy) if (market_cap and equity_fy not in (None, 0)) else y["pb"]
    dte = (debt_latest / equity_fy) if (debt_latest is not None and equity_fy not in (None, 0)) else None

    return {
        "ticker": yahoo_sym,
        "name": legal_name,
        "sector": sector,
        "industry": industry,
        "price": float(price) if price is not None else None,
        "shares": int(shares) if shares else None,
        "market_cap": float(market_cap) if market_cap is not None else None,
        "pe": float(pe) if pe is not None else None,
        "ps": float(ps) if ps is not None else None,
        "pb": float(pb) if pb is not None else None,
        "revenue_fy": revenue_fy,
        "net_income_fy": net_income_fy,
        "equity_fy": equity_fy,
        "debt": debt_latest,
        "debt_to_equity": dte,
        "rev_cagr_3y": rev_cagr_3y,
        "net_margin": net_margin,
        "updated_at": datetime.utcnow(),
    }


# ---------- JSON helpers ----------
def _df_to_json_records(df: pd.DataFrame):
    if df.empty:
        return []
    df = df.copy()
    if "updated_at" in df.columns:
        df["updated_at"] = df["updated_at"].astype(str)
    df = df.where(pd.notnull(df), None)
    return df.to_dict(orient="records")


# ---------- API: health ----------
@app.get("/health")
def health():
    return {"status": "ok"}


# ---------- METADATA for UI ----------
@app.get("/meta")
def meta():
    con: duckdb.DuckDBPyConnection = app.state.con
    sectors_metrics = [r[0] for r in con.execute(
        "SELECT DISTINCT sector FROM metrics WHERE sector IS NOT NULL ORDER BY 1").fetchall()]
    sectors_symbols = [r[0] for r in con.execute(
        "SELECT DISTINCT sector FROM symbols WHERE sector IS NOT NULL ORDER BY 1").fetchall()]
    sectors = sorted({*(s for s in sectors_metrics if s), *(s for s in sectors_symbols if s)})
    if not sectors:
        sectors = [
            "Technology", "Financial Services", "Consumer Cyclical", "Healthcare",
            "Industrials", "Energy", "Consumer Defensive", "Utilities",
            "Basic Materials", "Real Estate", "Communication Services",
        ]
    return {"sectors": sectors}


# ---------- UNIVERSE: add / list / count ----------
class SymbolIn(BaseModel):
    ticker: str
    name: Optional[str] = None
    exchange: Optional[str] = None
    country: Optional[str] = None
    currency: Optional[str] = None
    sector: Optional[str] = None
    industry: Optional[str] = None
    active: bool = True


@app.post("/universe/add")
def universe_add(symbols: List[SymbolIn]):
    con: duckdb.DuckDBPyConnection = app.state.con
    df = pd.DataFrame([s.dict() for s in symbols])
    if df.empty:
        return {"added": 0}
    df["ticker"] = df["ticker"].astype(str).str.strip()
    added = _upsert_df(con, "symbols", df)
    return {"added": added}


@app.get("/universe/list")
def universe_list(limit: int = 1000, offset: int = 0):
    con: duckdb.DuckDBPyConnection = app.state.con
    df = con.execute("SELECT * FROM symbols ORDER BY ticker LIMIT ? OFFSET ?", [int(limit), int(offset)]).df()
    return DefaultJSONResponse(_df_to_json_records(df))


@app.get("/universe/count")
def universe_count():
    con: duckdb.DuckDBPyConnection = app.state.con
    n = con.execute("SELECT COUNT(*) AS n FROM symbols WHERE coalesce(active, true)").fetchone()[0]
    return {"count": int(n)}


# ---------- UNIVERSE: bootstrap US via NASDAQ Trader ----------
@app.post("/universe/bootstrap_us")
def bootstrap_us():
    try:
        nasdaq_txt = _get_text("https://ftp.nasdaqtrader.com/dynamic/SymDir/nasdaqlisted.txt")
        other_txt = _get_text("https://ftp.nasdaqtrader.com/dynamic/SymDir/otherlisted.txt")
    except Exception as e:
        raise HTTPException(502, f"Failed to download NASDAQ symbol files: {e}")

    def read_table(txt: str) -> pd.DataFrame:
        lines = [ln for ln in txt.splitlines() if "File Creation Time" not in ln]
        csv_s = "\n".join(lines)
        return pd.read_csv(StringIO(csv_s), sep="|", dtype=str).fillna("")

    nasdaq = read_table(nasdaq_txt)
    other = read_table(other_txt)

    if "Test Issue" in nasdaq.columns:
        nasdaq = nasdaq[nasdaq["Test Issue"] != "Y"]
    if "Test Issue" in other.columns:
        other = other[other["Test Issue"] != "Y"]

    nasdaq_symbols = pd.DataFrame(
        {
            "ticker": nasdaq["Symbol"].str.strip(),
            "name": nasdaq["Security Name"].str.strip(),
            "exchange": "NASDAQ",
            "country": "US",
            "currency": "USD",
            "sector": None,
            "industry": None,
            "active": True,
        }
    )

    act = (
            (other["ACT Symbol"] if "ACT Symbol" in other.columns else None)
            or (other["Symbol"] if "Symbol" in other.columns else None)
            or (other["ACTSymbol"] if "ACTSymbol" in other.columns else None)
    )
    sec = (
            (other["Security Name"] if "Security Name" in other.columns else None)
            or (other["SecurityName"] if "SecurityName" in other.columns else None)
    )
    ex = other["Exchange"] if "Exchange" in other.columns else pd.Series(["OTHER"] * len(other))

    other_symbols = pd.DataFrame(
        {
            "ticker": act.fillna("").astype(str).str.strip(),
            "name": sec.fillna("").astype(str).str.strip(),
            "exchange": ex.fillna("").astype(str).str.strip(),
            "country": "US",
            "currency": "USD",
            "sector": None,
            "industry": None,
            "active": True,
        }
    )
    other_symbols = other_symbols[other_symbols["ticker"] != ""]

    universe = pd.concat([nasdaq_symbols, other_symbols], ignore_index=True).drop_duplicates(subset=["ticker"])
    con: duckdb.DuckDBPyConnection = app.state.con
    added = _upsert_df(con, "symbols", universe)
    return {"added": added}


# ---------- UNIVERSE: bootstrap global (best-effort public lists) ----------
@app.post("/universe/bootstrap_global")
def bootstrap_global(include_us: bool = True):
    sources: List[tuple] = [
        ("LSE",
         "https://raw.githubusercontent.com/datasets/london-stock-exchange-listed-companies/master/data/companies.csv",
         "UK", ".L"),
        ("TSX", "https://raw.githubusercontent.com/RobinL/freetickers/main/tsx.csv", "CA", ".TO"),
        ("ASX", "https://raw.githubusercontent.com/RobinL/freetickers/main/asx.csv", "AU", ".AX"),
        ("HKEX", "https://raw.githubusercontent.com/RobinL/freetickers/main/hk.csv", "HK", ".HK"),
        ("TSE", "https://raw.githubusercontent.com/RobinL/freetickers/main/tse.csv", "JP", ".T"),
    ]
    rows = []

    def _safe_fetch_csv(url: str) -> pd.DataFrame:
        try:
            txt = _get_text(url)
            df = pd.read_csv(StringIO(txt))
            return df
        except Exception:
            return pd.DataFrame()

    if include_us:
        try:
            added = bootstrap_us()
            rows.append(("US", added.get("added", 0)))
        except Exception:
            rows.append(("US", 0))

    con: duckdb.DuckDBPyConnection = app.state.con
    for ex, url, country, suffix in sources:
        df = _safe_fetch_csv(url)
        if df.empty:
            rows.append((ex, 0))
            continue
        cols = {c.lower().strip(): c for c in df.columns}
        sym_col = cols.get("symbol") or cols.get("ticker") or cols.get("code") or list(df.columns)[0]
        name_col = cols.get("name") or cols.get("company") or cols.get("security") or list(df.columns)[-1]
        sym = df[sym_col].astype(str).str.strip()
        nm = df[name_col].astype(str).str.strip()
        tickers = sym.apply(lambda s: s + suffix if suffix and not s.endswith(suffix) else s)
        out = pd.DataFrame(
            {
                "ticker": tickers,
                "name": nm,
                "exchange": ex,
                "country": country,
                "currency": None,
                "sector": df[cols.get("sector")] if "sector" in cols else None,
                "industry": df[cols.get("industry")] if "industry" in cols else None,
                "active": True,
            }
        ).dropna(subset=["ticker"]).drop_duplicates(subset=["ticker"])
        rows.append((ex, _upsert_df(con, "symbols", out)))
    return {"added_by_exchange": rows, "total_added": int(sum(n for _, n in rows))}


# ---------- INGEST ----------
@app.api_route("/ingest", methods=["POST", "GET"])
def ingest(
        tickers_body: Optional[List[str]] = Body(default=None, description="JSON array e.g. ['AAPL','MSFT']"),
        tickers_query: Optional[List[str]] = Query(default=None, alias="tickers",
                                                   description="Use multiple ?tickers= params"),
        parallel: int = Query(8, description="Max worker threads"),
):
    tickers = (tickers_body or tickers_query or [])
    if not tickers:
        raise HTTPException(400, "Provide tickers in JSON body or as ?tickers= query params")

    def worker(t):
        try:
            return compute_metrics_for(t)
        except Exception as e:
            return {"ticker": str(t).upper(), "error": str(e)}

    ok, err = [], []
    with ThreadPoolExecutor(max_workers=max(1, int(parallel))) as pool:
        for r in pool.map(worker, tickers):
            (ok if "error" not in r else err).append(r)
    if ok:
        df = pd.DataFrame(ok)
        con: duckdb.DuckDBPyConnection = app.state.con
        _upsert_df(con, "metrics", df)
    rows = ok + err
    for r in rows:
        if isinstance(r.get("updated_at"), datetime):
            r["updated_at"] = r["updated_at"].isoformat()
    return {"ingested": rows}


@app.post("/ingest_all")
def ingest_all(stale_hours: int = 24, limit: int = 5000, parallel: int = 8):
    con: duckdb.DuckDBPyConnection = app.state.con
    df = con.execute(
        """
        SELECT s.ticker, s.name, m.updated_at
        FROM symbols s
                 LEFT JOIN metrics m USING (ticker)
        WHERE coalesce(s.active, true)
        """
    ).df()
    cutoff = datetime.utcnow() - timedelta(hours=stale_hours)
    df["updated_at"] = pd.to_datetime(df["updated_at"])
    stale = df[(df["updated_at"].isna()) | (df["updated_at"] < cutoff)]
    tickers = stale["ticker"].head(int(limit)).tolist()
    if not tickers:
        return {"selected": 0, "ok": 0, "errors": 0}

    ok_rows, err_rows = [], []

    def worker(sym: str):
        try:
            return compute_metrics_for(sym)
        except Exception as e:
            return {"ticker": sym, "error": str(e)}

    with ThreadPoolExecutor(max_workers=max(1, int(parallel))) as pool:
        futures = [pool.submit(worker, t) for t in tickers]
        for fut in as_completed(futures):
            r = fut.result()
            (ok_rows if "error" not in r else err_rows).append(r)

    if ok_rows:
        con: duckdb.DuckDBPyConnection = app.state.con
        _upsert_df(con, "metrics", pd.DataFrame(ok_rows))
    return {"selected": len(tickers), "ok": len(ok_rows), "errors": len(err_rows), "error_samples": err_rows[:5]}


# ----- Screen models -----
class FilterRule(BaseModel):
    field: str
    op: Literal["<", "<=", ">", ">=", "=", "!=", "in", "not in", "between"]
    value: Any


class SortRule(BaseModel):
    field: str
    dir: Literal["asc", "desc"] = "asc"


class ScreenRequest(BaseModel):
    filters: List[FilterRule] = Field(default_factory=list)
    sort: List[SortRule] = Field(default_factory=list)
    limit: int = 100
    offset: int = 0


WHITELIST_FIELDS = {
    "ticker", "name", "sector", "industry", "price", "shares",
    "market_cap", "pe", "ps", "pb", "revenue_fy", "net_income_fy",
    "equity_fy", "debt", "debt_to_equity", "rev_cagr_3y", "net_margin", "updated_at",
}
FIELD_ALIASES = {
    "marketcap": "market_cap", "market_cap": "market_cap", "mktcap": "market_cap",
    "peratio": "pe", "p_e": "pe", "pe": "pe",
    "pricetosales": "ps", "price_to_sales": "ps", "ps": "ps",
    "pricetobook": "pb", "price_to_book": "pb", "pb": "pb",
    "revenue": "revenue_fy", "sales": "revenue_fy", "rev": "revenue_fy",
    "netincome": "net_income_fy", "net_income": "net_income_fy", "profit": "net_income_fy",
    "equity": "equity_fy", "book": "equity_fy",
    "dte": "debt_to_equity", "debttoequity": "debt_to_equity", "debt_to_equity": "debt_to_equity",
    "revcagr3y": "rev_cagr_3y", "rev_cagr_3y": "rev_cagr_3y",
    "netmargin": "net_margin", "net_margin": "net_margin",
    "price": "price", "shares": "shares", "ticker": "ticker", "name": "name",
    "sector": "sector", "industry": "industry", "debt": "debt", "updated_at": "updated_at",
}
WHITELIST_ORDER = {"asc", "desc"}


def _canon_field(raw: str) -> Optional[str]:
    key = "".join(ch for ch in raw.lower() if ch.isalnum() or ch == "_")
    return FIELD_ALIASES.get(key) or (raw if raw in WHITELIST_FIELDS else None)


# --- SCREEN: simple GET ---
@app.get("/screen")
def screen_simple(limit: int = 100, offset: int = 0):
    con: duckdb.DuckDBPyConnection = app.state.con
    df = con.execute("SELECT * FROM metrics LIMIT ? OFFSET ?", [int(limit), int(offset)]).df()
    return _df_to_json_records(df)


# --- SCREEN: POST with filters/sort ---
def _query_from_request(req: ScreenRequest):
    params: List[Any] = []

    def to_sql_condition(f: FilterRule) -> str:
        field = _canon_field(f.field)
        op = f.op.lower()
        val = f.value
        if not field:
            raise HTTPException(400, f"Bad field: {f.field}. Allowed: {sorted(WHITELIST_FIELDS)}")
        if op not in {"<", "<=", ">", ">=", "=", "!=", "in", "not in", "between"}:
            raise HTTPException(400, f"Bad operator: {f.op}")
        col = f'"{field}"'
        if op in {"<", "<=", ">", ">=", "=", "!="}:
            params.append(val)
            return f"{col} {op} ?"
        if op in {"in", "not in"}:
            if not isinstance(val, list) or len(val) == 0:
                raise HTTPException(400, "Value for IN/NOT IN must be a non-empty list")
            placeholders = ",".join(["?"] * len(val))
            params.extend(val)
            return f"{col} {'IN' if op == 'in' else 'NOT IN'} ({placeholders})"
        if not (isinstance(val, list) and len(val) == 2):
            raise HTTPException(400, "Value for BETWEEN must be [min, max]")
        params.extend([val[0], val[1]])
        return f"{col} BETWEEN ? AND ?"

    where_clause = " AND ".join(to_sql_condition(f) for f in req.filters) if req.filters else "TRUE"

    order_clause = ""
    if req.sort:
        order_terms = []
        for s in req.sort:
            field = _canon_field(s.field)
            direction = s.dir.lower()
            if not field:
                raise HTTPException(400, f"Bad sort field: {s.field}")
            if direction not in WHITELIST_ORDER:
                raise HTTPException(400, "Bad sort dir (use 'asc' or 'desc')")
            order_terms.append(f'"{field}" {direction}')
        if order_terms:
            order_clause = " ORDER BY " + ", ".join(order_terms)

    params.extend([int(req.limit), int(req.offset)])
    query = f"""
        SELECT * FROM metrics
        WHERE {where_clause}
        {order_clause}
        LIMIT ? OFFSET ?
    """
    return query, params


# Auto-hydration for screener (no manual ingest required)
def _ensure_metrics_for_candidates(tickers: List[str], ttl_hours: int = 24, parallel: int = 8):
    if not tickers:
        return
    con: duckdb.DuckDBPyConnection = app.state.con
    cutoff = datetime.utcnow() - timedelta(hours=int(ttl_hours))
    df = con.execute(
        "SELECT ticker, updated_at FROM metrics WHERE ticker IN (" + ",".join(["?"] * len(tickers)) + ")",
        tickers,
    ).df()
    stale = set(map(str.upper, tickers))
    for _, row in df.iterrows():
        try:
            if row["updated_at"] and pd.to_datetime(row["updated_at"]) >= cutoff:
                stale.discard(str(row["ticker"]).upper())
        except Exception:
            pass
    if not stale:
        return

    def worker(sym: str):
        try:
            return compute_metrics_for(sym)
        except Exception as e:
            return {"ticker": sym, "error": str(e)}

    ok_rows, err_rows = [], []
    with ThreadPoolExecutor(max_workers=max(1, int(parallel))) as pool:
        for r in pool.map(worker, list(stale)):
            (ok_rows if "error" not in r else err_rows).append(r)
    if ok_rows:
        _upsert_df(con, "metrics", pd.DataFrame(ok_rows))


@app.post("/screen")
def screen(req: ScreenRequest, ttl_hours: int = Query(24), auto_fill: bool = Query(True), parallel: int = Query(8)):
    con: duckdb.DuckDBPyConnection = app.state.con
    # Build candidate universe cheaply from symbols (use sector filter if provided)
    sectors_filter = [f for f in req.filters if _canon_field(f.field) == "sector" and f.op in ("in", "=")]
    where_sym, params = "coalesce(active, true)", []
    if sectors_filter:
        f = sectors_filter[0]
        if f.op == "in" and isinstance(f.value, list) and f.value:
            where_sym += " AND sector IN (" + ",".join(["?"] * len(f.value)) + ")"
            params.extend(f.value)
        elif f.op == "=":
            where_sym += " AND sector = ?"
            params.append(f.value)
    page_size = max(1, int(req.limit))
    cap = min(3000, page_size * 100)
    cand = con.execute(
        f"SELECT ticker FROM symbols WHERE {where_sym} LIMIT ?",
        [*params, cap],
    ).df()["ticker"].astype(str).str.upper().tolist()
    if auto_fill and cand:
        _ensure_metrics_for_candidates(cand, ttl_hours=ttl_hours, parallel=parallel)
    query, qparams = _query_from_request(req)
    df = con.execute(query, qparams).df()
    return _df_to_json_records(df)


# --- CSV export for current screen ---
@app.post("/screen.csv")
def screen_csv(req: ScreenRequest):
    query, params = _query_from_request(req)
    con: duckdb.DuckDBPyConnection = app.state.con
    df = con.execute(query, params).df()
    df = df.where(pd.notnull(df), None)
    if "updated_at" in df.columns:
        df["updated_at"] = df["updated_at"].astype(str)
    csv_bytes = df.to_csv(index=False).encode("utf-8")
    fname = f"screen_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    headers = {"Content-Disposition": f'attachment; filename="{fname}"', **NO_CACHE_HEADERS}
    return StreamingResponse(iter([csv_bytes]), media_type="text/csv", headers=headers)


# --------------------- FINANCIALS (QuickFS-like) ---------------------
def _fy_points(tag_obj: dict):
    out = []
    if not tag_obj or "units" not in tag_obj:
        return out
    for unit in ("USD", "USD/shares", "shares"):
        arr = tag_obj["units"].get(unit)
        if not arr:
            continue
        for x in arr:
            if x.get("fp") != "FY":
                continue
            val = x.get("val")
            if not isinstance(val, (int, float)):
                continue
            date = x.get("end") or x.get("instant")
            if not date or len(date) < 4:
                continue
            out.append((date[:4], float(val)))
        if out:
            break
    return out


def _first_fy_dict(facts: dict, tag_candidates: List[str]) -> dict:
    for t in tag_candidates:
        d = dict(_fy_points(facts.get(t, {})))
        if d:
            return d
    return {}


def _sum_year_dicts(*dicts: dict) -> dict:
    out = {}
    for d in dicts:
        for y, v in d.items():
            if isinstance(v, (int, float)):
                out[y] = out.get(y, 0.0) + v
    return out


def _sub_year_dicts(a: dict, b: dict) -> dict:
    out, years = {}, set(a) | set(b)
    for y in years:
        va, vb = a.get(y), b.get(y)
        if isinstance(va, (int, float)) and isinstance(vb, (int, float)):
            out[y] = va - vb
    return out


def _last_n_years(years: set[str], n: int = 10) -> list[str]:
    yrs = sorted({int(y) for y in years if str(y).isdigit()})
    return [str(y) for y in yrs[-n:]]


def _fmt_fin(v):
    if v is None:
        return "–"
    try:
        n = int(round(float(v)))
    except Exception:
        return "–"
    return f"({abs(n):,})" if n < 0 else f"{n:,}"


# ===== Ultra-expanded layouts with broad GAAP fallbacks =====
BALANCE_SHEET_LAYOUT = [
    {"kind": "header", "label": "Assets"},
    {"label": "Cash & Equivalents", "indent": 1, "tags": [
        "CashAndCashEquivalentsAtCarryingValue",
        "CashCashEquivalentsAndShortTermInvestments",
        "CashAndCashEquivalentsCarryingValue",
        "CashAndCashEquivalentsFairValueDisclosure",
        "Cash", "CashEquivalentsAtCarryingValue",
        "CashAndDueFromBanks",
        "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents"
    ]},
    {"label": "Short-Term Investments", "indent": 1, "tags": [
        "MarketableSecuritiesCurrent", "ShortTermInvestments",
        "ShortTermMarketableSecurities",
        "AvailableForSaleSecuritiesCurrent", "TradingSecuritiesCurrent",
        "DebtSecuritiesAvailableForSaleCurrent",
        "HeldToMaturitySecuritiesAmortizedCostCurrent",
        "OtherShortTermInvestments", "ShortTermInvestmentSecurities"
    ]},
    {"label": "Accounts Receivable, Net", "indent": 1, "tags": [
        "AccountsReceivableNetCurrent", "AccountsReceivableTradeNetCurrent",
        "ReceivablesNetCurrent", "LoansAndLeasesReceivableNetCurrent"
    ]},
    {"label": "Inventory", "indent": 1, "tags": [
        "InventoryNet", "Inventory",
        "InventoryFinishedGoods", "InventoryWorkInProcess", "InventoryRawMaterials"
    ]},
    {"label": "Prepaid & Other Current Assets", "indent": 1, "tags": [
        "PrepaidExpenseAndOtherCurrentAssets", "OtherAssetsCurrent",
        "PrepaidExpenseCurrent"
    ]},
    {"label": "Deferred Tax Assets (Current)", "indent": 1, "tags": [
        "DeferredTaxAssetsNetCurrent", "DeferredIncomeTaxAssetsCurrent"
    ]},
    {"label": "Total Current Assets", "indent": 0, "tags": ["AssetsCurrent", "CurrentAssets"], "total_of": [
        "Cash & Equivalents", "Short-Term Investments", "Accounts Receivable, Net",
        "Inventory", "Prepaid & Other Current Assets", "Deferred Tax Assets (Current)"
    ], "is_total": True},
    {"label": "Long-Term Investments", "indent": 0, "tags": [
        "MarketableSecuritiesNoncurrent", "AvailableForSaleSecuritiesNoncurrent",
        "LongTermInvestments", "EquityMethodInvestments",
        "HeldToMaturitySecuritiesNoncurrent"
    ]},
    {"label": "Property, Plant, & Equipment (Net)", "indent": 0, "tags": [
        "PropertyPlantAndEquipmentNet",
        "PropertyPlantAndEquipmentIncludingConstructionInProgressNet",
        "PropertyPlantAndEquipmentExcludingFinancialLeasesNet",
        "PropertyPlantAndEquipmentNetOfAccumulatedDepreciationAndAmortization"
    ]},
    {"label": "Operating Lease ROU Assets", "indent": 0, "tags": ["OperatingLeaseRightOfUseAsset"]},
    {"label": "Intangible Assets (Net)", "indent": 0, "tags": [
        "IntangibleAssetsNetExcludingGoodwill", "FiniteLivedIntangibleAssetsNet",
        "IndefiniteLivedIntangibleAssetsExcludingGoodwill",
        "IntangibleAssetsNet"
    ]},
    {"label": "Goodwill", "indent": 0, "tags": [
        "Goodwill", "GoodwillAndIntangibleAssetsNet", "GoodwillAndOtherIntangibleAssetsNet"
    ]},
    {"label": "Deferred Tax Assets (Noncurrent)", "indent": 0, "tags": [
        "DeferredTaxAssetsNetNoncurrent", "DeferredIncomeTaxAssetsNet"
    ]},
    {"label": "Other Assets", "indent": 0, "tags": [
        "OtherAssetsNoncurrent", "OtherAssets", "OtherNoncurrentAssets", "OtherAssetsMiscellaneous"
    ]},
    {"label": "Total Assets", "indent": 0, "tags": ["Assets"], "total_of": [
        "Total Current Assets", "Long-Term Investments", "Property, Plant, & Equipment (Net)",
        "Operating Lease ROU Assets", "Intangible Assets (Net)", "Goodwill",
        "Deferred Tax Assets (Noncurrent)", "Other Assets"
    ], "is_total": True},

    {"kind": "header", "label": "Liabilities & Equity"},
    {"label": "Short-Term Debt", "indent": 1, "tags": [
        "DebtCurrent", "ShortTermBorrowings", "CommercialPaper", "ShortTermDebt",
        "LongTermDebtCurrent", "CurrentPortionOfLongTermDebtAndCapitalLeaseObligations"
    ]},
    {"label": "Accounts Payable", "indent": 1, "tags": [
        "AccountsPayableCurrent", "TradeAndOtherPayablesCurrent"
    ]},
    {"label": "Accrued Liabilities", "indent": 1, "tags": [
        "AccruedLiabilitiesCurrent", "OtherLiabilitiesCurrent",
        "AccruedExpensesCurrent"
    ]},
    {"label": "Deferred Revenue (Current)", "indent": 1, "tags": [
        "ContractWithCustomerLiabilityCurrent", "DeferredRevenueAndCreditsCurrent",
        "UnearnedRevenueCurrent"
    ]},
    {"label": "Operating Lease Liabilities (Current)", "indent": 1, "tags": [
        "OperatingLeaseLiabilityCurrent"
    ]},
    {"label": "Income Taxes Payable (Current)", "indent": 1, "tags": [
        "TaxesPayableCurrent", "IncomeTaxesPayableCurrent"
    ]},
    {"label": "Total Current Liabilities", "indent": 0, "tags": ["LiabilitiesCurrent"], "total_of": [
        "Short-Term Debt", "Accounts Payable", "Accrued Liabilities",
        "Deferred Revenue (Current)", "Operating Lease Liabilities (Current)",
        "Income Taxes Payable (Current)"
    ], "is_total": True},

    {"label": "Long-Term Debt", "indent": 0, "tags": [
        "LongTermDebtNoncurrent", "LongTermDebt", "DebtNoncurrent",
        "LongTermBorrowings", "LongTermDebtAndCapitalLeaseObligations"
    ]},
    {"label": "Deferred Revenue (Noncurrent)", "indent": 0, "tags": [
        "ContractWithCustomerLiabilityNoncurrent", "DeferredRevenueAndCreditsNoncurrent"
    ]},
    {"label": "Operating Lease Liabilities (Noncurrent)", "indent": 0, "tags": ["OperatingLeaseLiabilityNoncurrent"]},
    {"label": "Deferred Tax Liabilities (Noncurrent)", "indent": 0, "tags": [
        "DeferredTaxLiabilitiesNoncurrent", "DeferredIncomeTaxLiabilitiesNet"
    ]},
    {"label": "Other Liabilities (Noncurrent)", "indent": 0, "tags": ["OtherLiabilitiesNoncurrent"]},
    {"label": "Total Liabilities", "indent": 0, "tags": ["Liabilities", "LiabilitiesCurrentAndNoncurrent"],
     "is_total": True},

    {"label": "Common Stock & APIC", "indent": 0, "tags": [
        "CommonStockAndAdditionalPaidInCapital", "AdditionalPaidInCapital",
        "CommonStockValue", "CapitalStockValue"
    ]},
    {"label": "Treasury Stock", "indent": 0, "tags": ["TreasuryStockValue", "TreasuryStockCommonValue"]},
    {"label": "Retained Earnings (Accumulated Deficit)", "indent": 0, "tags": ["RetainedEarningsAccumulatedDeficit"]},
    {"label": "AOCI", "indent": 0, "tags": [
        "AccumulatedOtherComprehensiveIncomeLossNetOfTax",
        "AccumulatedOtherComprehensiveIncomeLoss"
    ]},
    {"label": "Noncontrolling Interests", "indent": 0, "tags": ["MinorityInterest", "TemporaryEquityCarryingAmount"]},
    {"label": "Total Equity", "indent": 0, "tags": [
        "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest",
        "StockholdersEquity", "Equity", "CommonStockholdersEquity",
        "MembersEquity", "PartnersCapital"
    ], "is_total": True},
    {"label": "Total Liabilities & Equity", "indent": 0, "tags": [
        "LiabilitiesAndStockholdersEquity", "LiabilitiesAndEquity", "LiabilitiesAndPartnersCapital"
    ], "total_of": ["Total Liabilities", "Total Equity"], "is_total": True},
]
INCOME_STATEMENT_LAYOUT = [
    {"label": "Revenue", "indent": 0, "tags": [
        "RevenueFromContractWithCustomerExcludingAssessedTax",
        "Revenues", "Revenue", "SalesRevenueNet",
        "SalesRevenueGoodsNet", "SalesRevenueServicesNet", "TotalRevenue"
    ], "is_total": True},
    {"label": "Interest Income", "indent": 1,
     "tags": ["InterestAndDividendIncomeOperating", "InterestIncome", "InterestAndDividendIncomeSecurities"]},
    {"label": "Interest Expense", "indent": 1,
     "tags": ["InterestExpense", "InterestAndDebtExpense", "InterestExpenseDebt"]},
    {"label": "Cost of Goods Sold", "indent": 0, "tags": [
        "CostOfRevenue", "CostOfGoodsSold", "CostOfSales",
        "CostOfGoodsAndServicesSold", "CostOfGoodsSoldExcludingDepreciationDepletionAndAmortization"
    ]},
    {"label": "Gross Profit", "indent": 0, "tags": ["GrossProfit", "GrossProfitLoss"],
     "calc": {"type": "subtract", "from": "Revenue", "minus": ["Cost of Goods Sold"]}, "is_total": True},
    {"label": "Research & Development", "indent": 1,
     "tags": ["ResearchAndDevelopmentExpense", "ResearchAndDevelopment"]},
    {"label": "Selling, General & Administrative", "indent": 1,
     "tags": ["SellingGeneralAndAdministrativeExpense", "SellingAndMarketingExpense",
              "GeneralAndAdministrativeExpense"]},
    {"label": "Stock-Based Compensation", "indent": 1, "tags": ["ShareBasedCompensation"]},
    {"label": "Restructuring & Other", "indent": 1,
     "tags": ["RestructuringCharges", "RestructuringAndMergerAndAcquisition"]},
    {"label": "Other Operating Income (Expense)", "indent": 1, "tags": ["OtherOperatingIncomeExpenseNet"]},
    {"label": "Operating Profit", "indent": 0, "tags": ["OperatingIncomeLoss", "IncomeFromOperations"],
     "is_total": True},
    {"label": "Nonoperating Income (Expense)", "indent": 0,
     "tags": ["NonoperatingIncomeExpense", "OtherNonoperatingIncomeExpense"]},
    {"label": "Pre-Tax Income", "indent": 0, "tags": [
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest",
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesMinorityInterestAndIncomeLossFromEquityMethodInvestments",
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxes", "PretaxIncomeLoss"
    ], "is_total": True},
    {"label": "Income Tax", "indent": 0,
     "tags": ["IncomeTaxExpenseBenefit", "IncomeTaxExpenseBenefitContinuingOperations", "IncomeTaxExpense"]},
    {"label": "Net Income", "indent": 0, "tags": [
        "NetIncomeLoss", "ProfitLoss",
        "NetIncomeLossAttributableToParent",
        "NetIncomeLossIncludingPortionAttributableToNoncontrollingInterest",
        "NetIncomeLossAvailableToCommonStockholdersBasic",
        "NetIncomeLossAvailableToCommonStockholdersDiluted"
    ], "is_total": True},
]
CASH_FLOW_LAYOUT = [
    {"label": "Net Income", "indent": 0, "tags": [
        "NetIncomeLoss", "ProfitLoss",
        "NetIncomeLossAttributableToParent",
        "NetIncomeLossIncludingPortionAttributableToNoncontrollingInterest"
    ], "is_total": True},
    {"label": "Depreciation & Amortization", "indent": 1, "tags": [
        "DepreciationDepletionAndAmortization",
        "DepreciationAmortizationAndAccretionNet",
        "DepreciationAmortizationAndAccretion",
        "DepreciationAndAmortization",
        "AmortizationOfIntangibleAssets", "Depreciation"
    ]},
    {"label": "Stock-Based Compensation", "indent": 1, "tags": ["ShareBasedCompensation"]},
    {"label": "Deferred Income Taxes", "indent": 1,
     "tags": ["DeferredIncomeTaxExpenseBenefit", "DeferredIncomeTaxesAndTaxCredits"]},
    {"label": "Other Non-Cash Items", "indent": 1,
     "tags": ["OtherNoncashIncomeExpense", "OtherNoncashOperatingActivities"]},
    {"label": "Change in Accounts Receivable", "indent": 1, "tags": ["IncreaseDecreaseInAccountsReceivable"]},
    {"label": "Change in Inventory", "indent": 1, "tags": ["IncreaseDecreaseInInventories"]},
    {"label": "Change in Accounts Payable", "indent": 1,
     "tags": ["IncreaseDecreaseInAccountsPayableTrade", "IncreaseDecreaseInAccountsPayable"]},
    {"label": "Change in Accrued Liabilities", "indent": 1,
     "tags": ["IncreaseDecreaseInAccruedLiabilities", "IncreaseDecreaseInOtherCurrentLiabilities"]},
    {"label": "Change in Deferred Revenue", "indent": 1,
     "tags": ["IncreaseDecreaseInContractWithCustomerLiability", "IncreaseDecreaseInDeferredRevenue"]},
    {"label": "Change in Other Operating Assets/Liabilities", "indent": 1,
     "tags": ["IncreaseDecreaseInOtherOperatingAssets", "IncreaseDecreaseInOtherOperatingLiabilities"]},
    {"label": "Cash From Operations", "indent": 0, "tags": [
        "NetCashProvidedByUsedInOperatingActivities",
        "NetCashProvidedByUsedInOperatingActivitiesContinuingOperations",
        "NetCashProvidedByUsedInOperatingActivitiesIndirect",
        "NetCashProvidedByUsedInOperatingActivitiesDirect"
    ], "is_total": True},
    {"label": "Property, Plant, & Equipment (Capex)", "indent": 1, "tags": [
        "PaymentsToAcquirePropertyPlantAndEquipment",
        "CapitalExpenditures",
        "PaymentsToAcquireProductiveAssets",
        "PaymentsToAcquirePropertyPlantAndEquipmentClassifiedAsInvestingActivities"
    ]},
    {"label": "Proceeds from Sale of PP&E", "indent": 1, "tags": ["ProceedsFromSaleOfPropertyPlantAndEquipment"]},
    {"label": "Acquisitions (Net of Cash)", "indent": 1, "tags": [
        "PaymentsToAcquireBusinessesNetOfCashAcquired",
        "PaymentsToAcquireBusinessesAndInterestInAffiliatesNetOfCashAcquired"
    ]},
    {"label": "Purchases of Investments", "indent": 1, "tags": [
        "PaymentsToAcquireInvestments",
        "PaymentsToAcquireAvailableForSaleSecuritiesDebt",
        "PaymentsToAcquireHeldToMaturitySecurities"
    ]},
    {"label": "Proceeds from Investments", "indent": 1, "tags": [
        "ProceedsFromSaleOfInvestments",
        "ProceedsFromMaturitiesPrepaymentsAndCallsOfAvailableForSaleSecurities",
        "ProceedsFromSaleOfAvailableForSaleSecuritiesDebt"
    ]},
    {"label": "Cash From Investing", "indent": 0, "tags": [
        "NetCashProvidedByUsedInInvestingActivities",
        "NetCashProvidedByUsedInInvestingActivitiesContinuingOperations"
    ], "is_total": True},
    {"label": "Dividends Paid", "indent": 1, "tags": ["PaymentsOfDividends", "PaymentsOfDividendsCommonStock"]},
    {"label": "Share Repurchases", "indent": 1,
     "tags": ["PaymentsForRepurchaseOfCommonStock", "PaymentsForRepurchaseOfEquity"]},
    {"label": "Proceeds from Stock Issuance", "indent": 1,
     "tags": ["ProceedsFromIssuanceOfCommonStock", "ProceedsFromStockOptionsExercised"]},
    {"label": "Proceeds from Issuance of Debt", "indent": 1,
     "tags": ["ProceedsFromIssuanceOfLongTermDebt", "ProceedsFromBorrowings", "DebtIssued"]},
    {"label": "Repayments of Debt", "indent": 1,
     "tags": ["RepaymentsOfLongTermDebt", "RepaymentsOfDebt", "DebtRepayments"]},
    {"label": "Cash From Financing", "indent": 0, "tags": [
        "NetCashProvidedByUsedInFinancingActivities",
        "NetCashProvidedByUsedInFinancingActivitiesContinuingOperations"
    ], "is_total": True},
    {"label": "Free Cash Flow", "indent": 0, "tags": [],
     "calc": {"type": "subtract", "from": "Cash From Operations", "minus": ["Property, Plant, & Equipment (Capex)"]},
     "is_total": True},
]
# ---------- Yahoo fallback for financial tables (Improved) ----------
import re
from typing import Literal


def _yf_get_df(t: yf.Ticker, kind: Literal["bs", "is", "cf"],
               freq: Literal["yearly", "quarterly"] = "yearly") -> pd.DataFrame:
    """
    Get Yahoo Finance balance sheet, income statement, or cash flow.
    Tries the given frequency, then an alternate naming if available.
    """
    getter = {
        "bs": t.get_balance_sheet,
        "is": t.get_income_stmt,
        "cf": t.get_cashflow
    }[kind]
    for f in (freq, "annual" if freq == "yearly" else "quarterly"):
        try:
            df = getter(freq=f)
            if isinstance(df, pd.DataFrame) and not df.empty:
                return df
        except Exception:
            pass
    return pd.DataFrame()


def _yf_annual_series(df: pd.DataFrame, key: str) -> dict:
    """
    Retrieve values for a row label (case/space-insensitive) and extract years.
    Works with Period, Timestamp, or string-based columns.
    """
    if not isinstance(df, pd.DataFrame) or df.empty:
        return {}
    # Normalize index
    norm_index = {str(ix).strip().lower().replace(" ", ""): ix for ix in df.index}
    k = key.strip().lower().replace(" ", "")
    row_key = norm_index.get(k)
    if row_key is None:
        # Try partial match
        row_key = next((orig for nk, orig in norm_index.items() if k in nk), None)
    if row_key is None:
        return {}
    try:
        s = df.loc[row_key].dropna()
    except Exception:
        return {}
    out = {}
    for col, val in s.items():
        year = None
        try:
            if hasattr(col, "year"):
                year = int(getattr(col, "year"))
            else:
                m = re.search(r"(\d{4})", str(col))
                if m:
                    year = int(m.group(1))
        except Exception:
            year = None
        if year:
            try:
                out[str(year)] = float(val)
            except Exception:
                pass
    return out


# --- aliases unchanged (kept intact) ---
YF_ALIASES = {
    # Balance sheet
    "Cash & Equivalents": ["Cash And Cash Equivalents", "Cash And Cash Equivalents And Short Term Investments", "Cash",
                           "Cash And Due From Banks"],
    "Short-Term Investments": ["Short Term Investments", "Marketable Securities", "Other Short Term Investments"],
    "Accounts Receivable, Net": ["Net Receivables", "Accounts Receivable"],
    "Inventory": ["Inventory"],
    "Prepaid & Other Current Assets": ["Other Current Assets", "Prepaid Expenses"],
    "Deferred Tax Assets (Current)": ["Deferred Tax Assets"],
    "Total Current Assets": ["Total Current Assets", "Current Assets Total"],
    "Long-Term Investments": ["Long Term Investments"],
    "Property, Plant, & Equipment (Net)": ["Property Plant Equipment", "Property Plant Equipment Net",
                                           "Net Property Plant And Equipment"],
    "Operating Lease ROU Assets": ["Operating Lease Right Of Use Asset"],
    "Intangible Assets (Net)": ["Intangible Assets", "Intangible Assets Net"],
    "Goodwill": ["Goodwill", "Goodwill Net"],
    "Deferred Tax Assets (Noncurrent)": ["Deferred Long Term Asset Charges", "Deferred Tax Assets"],
    "Other Assets": ["Other Non Current Assets", "Other Noncurrent Assets", "Other Assets"],
    "Total Assets": ["Total Assets", "Assets"],
    "Short-Term Debt": ["Short Long Term Debt", "Short Term Debt", "Commercial Paper",
                        "Current Portion Of Long Term Debt"],
    "Accounts Payable": ["Accounts Payable"],
    "Accrued Liabilities": ["Other Current Liabilities", "Accrued Expenses"],
    "Deferred Revenue (Current)": ["Deferred Revenue"],
    "Operating Lease Liabilities (Current)": ["Operating Lease Liability Current"],
    "Income Taxes Payable (Current)": ["Income Taxes Payable"],
    "Total Current Liabilities": ["Total Current Liabilities"],
    "Long-Term Debt": ["Long Term Debt", "Long Term Borrowings", "Long Term Debt Noncurrent"],
    "Deferred Revenue (Noncurrent)": ["Deferred Long Term Liability Charges", "Deferred Revenue"],
    "Operating Lease Liabilities (Noncurrent)": ["Operating Lease Liability Non Current"],
    "Deferred Tax Liabilities (Noncurrent)": ["Deferred Long Term Liability Charges", "Deferred Tax Liabilities"],
    "Other Liabilities (Noncurrent)": ["Other Liabilities"],
    "Total Liabilities": ["Total Liab", "Total Liabilities Net Minority Interest", "Total Liabilities"],
    "Common Stock & APIC": ["Common Stock", "Capital Stock", "Additional Paid In Capital", "Capital Surplus"],
    "Treasury Stock": ["Treasury Stock"],
    "Retained Earnings (Accumulated Deficit)": ["Retained Earnings"],
    "AOCI": ["Accumulated Other Comprehensive Income"],
    "Noncontrolling Interests": ["Minority Interest"],
    "Total Equity": ["Total Stockholder Equity", "Stockholders Equity", "Total Equity"],
    "Total Liabilities & Equity": ["Total Liabilities And Stockholder Equity"],
    # Income statement
    "Revenue": ["Total Revenue", "Revenue", "Sales"],
    "Interest Income": ["Interest Income"],
    "Interest Expense": ["Interest Expense"],
    "Cost of Goods Sold": ["Cost Of Revenue", "Cost Of Goods And Services Sold", "Cost Of Sales"],
    "Gross Profit": ["Gross Profit", "Gross Profit Loss"],
    "Research & Development": ["Research Development"],
    "Selling, General & Administrative": ["Selling General Administrative", "Selling/General/Admin. Expenses"],
    "Stock-Based Compensation": ["Stock Based Compensation"],
    "Restructuring & Other": ["Restructuring And Merger And Acquisition"],
    "Other Operating Income (Expense)": ["Operating Income Others", "Other Operating Expenses"],
    "Operating Profit": ["Operating Income", "Operating Income Or Loss"],
    "Nonoperating Income (Expense)": ["Total Other Income Expense Net"],
    "Pre-Tax Income": ["Income Before Tax"],
    "Income Tax": ["Income Tax Expense", "Income Tax Expense Benefit"],
    "Net Income": ["Net Income Common Stockholders", "Net Income", "Net Income Applicable To Common Shares"],
    # Cash flow
    "Depreciation & Amortization": ["Depreciation", "Depreciation And Amortization", "Amortization"],
    "Stock-Based Compensation": ["Stock Based Compensation"],
    "Deferred Income Taxes": ["Deferred Income Tax"],
    "Other Non-Cash Items": ["Other Non Cash Items"],
    "Change in Accounts Receivable": ["Change In Accounts Receivable"],
    "Change in Inventory": ["Change In Inventory"],
    "Change in Accounts Payable": ["Change In Accounts Payable"],
    "Change in Accrued Liabilities": ["Change In Other Current Liabilities"],
    "Change in Deferred Revenue": ["Change In Deferred Revenue"],
    "Change in Other Operating Assets/Liabilities": ["Change In Other Current Assets",
                                                     "Change In Other Working Capital"],
    "Cash From Operations": ["Total Cash From Operating Activities", "Net Cash Provided By Operating Activities"],
    "Property, Plant, & Equipment (Capex)": ["Capital Expenditures", "Purchase Of Property Plant And Equipment"],
    "Proceeds from Sale of PP&E": ["Sale Of Property Plant Equipment"],
    "Acquisitions (Net of Cash)": ["Net Business Purchase And Sale"],
    "Purchases of Investments": ["Purchase Of Investments", "Net Investment Purchase And Sale"],
    "Proceeds from Investments": ["Sale Of Investments"],
    "Cash From Investing": ["Total Cashflows From Investing Activities",
                            "Net Cash Provided By Used In Investing Activities"],
    "Dividends Paid": ["Dividends Paid"],
    "Share Repurchases": ["Repurchase Of Stock"],
    "Proceeds from Stock Issuance": ["Common Stock Issued", "Proceeds From Stock Options Exercised"],
    "Proceeds from Issuance of Debt": ["Net Borrowings", "Debt Issued"],
    "Repayments of Debt": ["Debt Repayment"],
    "Cash From Financing": ["Total Cash From Financing Activities",
                            "Net Cash Provided By Used In Financing Activities"],
    "Free Cash Flow": ["Free Cash Flow"],
}


def _order_period_labels(labels, order: Literal["asc", "desc"]="desc"):
    import re
    def key(x):
        s = str(x)
        m = re.match(r"^(\d{4})(?:[ -_/]?Q([1-4]))?$", s, re.I)
        if m:
            return (int(m.group(1)), int(m.group(2) or 0), "")
        return (10**9, 0, s)
    out = sorted(labels, key=key)
    return [str(x) for x in (out if order=="asc" else reversed(out))]

def _layout_to_rows(layout: List[dict], years: List[str], values_by_label: dict) -> List[dict]:
    rows = []
    for item in layout:
        kind = item.get("kind")
        if kind in ("header", "subheader"):
            rows.append({"banner": kind, "label": item["label"]})
            continue
        if item.get("hidden"):
            continue

        label = item.get("label", "")
        vals = values_by_label.get(label, {})
        has_any = any(vals.get(y) is not None for y in years)

        if not has_any and not item.get("is_total"):
            continue

        rows.append({
            "banner": None,
            "label": label,
            "indent": item.get("indent", 0),
            "is_total": item.get("is_total", False),
            "values": [vals.get(y) for y in years],
        })
    return rows

# ---------- Yahoo builder ----------

def _build_matrix_yahoo(
    ticker: str,
    layout: List[dict],
    freq: Literal["yearly", "quarterly"]="yearly",
    order: Literal["asc","desc"]="desc",
):
    try:
        t = yf.Ticker(ticker)
        bs = _yf_get_df(t, "bs", freq)
        is_ = _yf_get_df(t, "is", freq)
        cf = _yf_get_df(t, "cf", freq)
    except Exception:
        bs = is_ = cf = pd.DataFrame()

    def series_for_label(lbl: str) -> dict:
        for df in (is_, bs, cf):
            if not isinstance(df, pd.DataFrame) or df.empty:
                continue
            for key in YF_ALIASES.get(lbl, []):
                d = _yf_annual_series(df, key)
                if d:
                    return d
        return {}

    values_by_label, all_periods = {}, set()
    for item in layout:
        if item.get("kind") in ("header", "subheader") or item.get("hidden"):
            continue
        d = series_for_label(item["label"])
        values_by_label[item["label"]] = d
        all_periods |= set(d.keys())

    n_cols = 10 if freq == "yearly" else 8
    years = _last_n_years(all_periods, n_cols)
    years = _order_period_labels(years, order)

    def _sum_for(labels):
        return _sum_year_dicts(*[values_by_label.get(lbl, {}) for lbl in labels])

    # totals & calcs
    for item in layout:
        if item.get("kind") in ("header", "subheader") or item.get("hidden"):
            continue
        label = item["label"]
        reported = values_by_label.get(label, {})

        if item.get("total_of"):
            computed = _sum_for(item["total_of"])
            values_by_label[label] = {y: (reported.get(y) if reported.get(y) is not None else computed.get(y))
                                      for y in years}

        calc = item.get("calc")
        if calc:
            typ = calc.get("type")
            if typ == "subtract":
                minuend = values_by_label.get(calc["from"], {})
                minus_labels = calc.get("minus", [])
                if isinstance(minus_labels, str):
                    minus_labels = [minus_labels]
                sub = _sum_year_dicts(*[values_by_label.get(lbl, {}) for lbl in minus_labels])
                computed = _sub_year_dicts(minuend, sub)
            elif typ == "sum":
                computed = _sum_year_dicts(*[values_by_label.get(lbl, {}) for lbl in calc.get("items", [])])
            else:
                computed = {}
            values_by_label[label] = {y: (reported.get(y) if reported.get(y) is not None else computed.get(y))
                                      for y in years}

    rows = _layout_to_rows(layout, years, values_by_label)
    meta = {"ticker": ticker.upper(), "cik": "-", "years": years, "source": f"Yahoo ({freq})"}
    return rows, meta

# ---------- SEC (annual) ----------

def _build_matrix_10y(ticker: str, layout: List[dict], order: Literal["asc","desc"]="desc"):
    cmap = _company_map()
    sec_key = _sec_lookup_key(ticker)
    if sec_key.upper() not in cmap:
        raise HTTPException(404, f"Financials require a US SEC filer. Not found in SEC list: {ticker.upper()}")
    cik, _ = cmap[sec_key.upper()]
    facts = _companyfacts(cik).get("facts", {}).get("us-gaap", {})

    values_by_label, all_years = {}, set()
    for item in layout:
        if item.get("kind") in ("header", "subheader") or item.get("hidden"):
            continue
        vals = _first_fy_dict(facts, item.get("tags", []))
        values_by_label[item["label"]] = vals
        all_years |= set(vals.keys())

    years = _order_period_labels(_last_n_years(all_years, 10), order)

    def _sum_for(labels):
        return _sum_year_dicts(*[values_by_label.get(lbl, {}) for lbl in labels])

    for item in layout:
        if item.get("kind") in ("header", "subheader") or item.get("hidden"):
            continue
        label = item["label"]
        reported = values_by_label.get(label, {})

        if item.get("total_of"):
            computed = _sum_for(item["total_of"])
            values_by_label[label] = {y: (reported.get(y) if reported.get(y) is not None else computed.get(y))
                                      for y in years}

        calc = item.get("calc")
        if calc:
            typ = calc.get("type")
            if typ == "subtract":
                minuend = values_by_label.get(calc["from"], {})
                minus_labels = calc.get("minus", [])
                if isinstance(minus_labels, str):
                    minus_labels = [minus_labels]
                sub = _sum_year_dicts(*[values_by_label.get(lbl, {}) for lbl in minus_labels])
                computed = _sub_year_dicts(minuend, sub)
            elif typ == "sum":
                computed = _sum_year_dicts(*[values_by_label.get(lbl, {}) for lbl in calc.get("items", [])])
            else:
                computed = {}
            values_by_label[label] = {y: (reported.get(y) if reported.get(y) is not None else computed.get(y))
                                      for y in years}

    rows = _layout_to_rows(layout, years, values_by_label)
    meta = {"ticker": ticker.upper(), "cik": cik, "years": years, "source": "SEC"}
    return rows, meta

# ---------- Unified ----------

def _build_matrix_any(
    ticker: str,
    layout: List[dict],
    *,
    freq: Literal["annual","quarterly"]="annual",
    order: Literal["asc","desc"]="desc",
):
    if freq == "annual":
        try:
            rows, meta = _build_matrix_10y(ticker, layout, order=order)
            meta["source"] = "SEC (annual)"
            return rows, meta
        except Exception as e:
            print(f"SEC annual failed for {ticker}: {e}. Falling back to Yahoo.")
    yf_freq = "yearly" if freq == "annual" else "quarterly"
    return _build_matrix_yahoo(ticker, layout, freq=yf_freq, order=order)

# ---------- Handler (stop right before CSS) ----------

@app.get("/financials/{ticker}/structured.html", response_class=HTMLResponse)
def structured_html(
    ticker: str, request: Request,
    freq: Literal["annual","quarterly"]="annual",
    order: Literal["asc","desc"]="desc",
):
    ticker = _normalize_ticker(ticker)

    bs_rows, meta_bs = _build_matrix_any(ticker, BALANCE_SHEET_LAYOUT, freq=freq, order=order)
    is_rows, meta_is = _build_matrix_any(ticker, INCOME_STATEMENT_LAYOUT, freq=freq, order=order)
    cf_rows, meta_cf = _build_matrix_any(ticker, CASH_FLOW_LAYOUT, freq=freq, order=order)

    years = meta_bs["years"]  # all three share the same axis by construction
    # Units scaler (billions/millions/thousands); default to "millions"
    units_param = (request.query_params.get("units") or "millions").strip().lower()
    _SCALE = {"billions": 1_000_000_000, "millions": 1_000_000, "thousands": 1_000}
    _scale = _SCALE.get(units_param, 1_000_000)

    def _fmt_scaled(v):
        if v is None:
            return "–"
        try:
            return _fmt_fin(float(v) / _scale)
        except Exception:
            return _fmt_fin(v)
    def render_table(title: str, rows: list, years: list[str]) -> str:
        th_years = "".join(f"<th class='year'>{y}</th>" for y in years)
        body = []
        for r in rows:
            if r.get("banner") == "header":
                body.append(f"<tr class='section'><td class='label section-title' colspan='{1+len(years)}'>{r['label']}</td></tr>")
                continue
            if r.get("banner") == "subheader":
                body.append(f"<tr class='subsection'><td class='label' colspan='{1+len(years)}'>{r['label']}</td></tr>")
                continue
            indent_px = 18 * r.get("indent", 0)
            row_cls = "total" if r.get("is_total") else ""
            tds = [f"<td class='label' style='padding-left:{indent_px}px'>{r['label']}</td>"]
            tds += [f"<td class='val'>{_fmt_scaled(v)}</td>" for v in r["values"]]
            body.append(f"<tr class='{row_cls}'>{''.join(tds)}</tr>")
        colgroup = "<col class='c-label'/>" + "".join("<col class='c-year'/>" for _ in years)
        return f"""
        <h2 class="title">{title}</h2>
        <table class="fin-table">
          {colgroup}
          <thead><tr><th class="label-h">Line</th>{th_years}</tr></thead>
          <tbody>{''.join(body)}</tbody>
        </table>
        """

    css = """
    <style>
  /* ===== Quiet-lux light theme (unchanged palette) ===== */
  :root{
    --bg:#f6f5f2; --card:#ffffff; --ink:#111111; --muted:#6e6e73; --line:#e6e2d9;
    --band:#fbfaf7; --elev:0 12px 32px rgba(0,0,0,.06); --accent:#c5a572; --focus:0 0 0 3px rgba(197,165,114,.25);
  }
  html,body{
    background:var(--bg); color:var(--ink);
    font-family:ui-sans-serif,-apple-system,BlinkMacSystemFont,"SF Pro Text",Inter,Segoe UI,Roboto,Arial,sans-serif;
    font-size:15px; letter-spacing:.01em; margin:0; height:100%;
  }
  /* ======= LAYOUT: full-width grid, proportional screener ======= */
  /* The screener is a responsive column: never <360px, scales with viewport, capped for readability */
  .grid{
  display:grid;
  grid-template-columns: minmax(0, 1fr) clamp(360px, 26vw, 560px); /* left grows, right is a responsive column */
  gap:16px;
  width:100%;
  max-width:100vw;            /* remove 1200px cap */
  margin:14px auto;
  padding:0 14px;
  align-items:start;
  box-sizing:border-box;
}

/* keep (or add) sticky right column */
.screener{ position: sticky; top:14px; height: calc(100vh - 28px); overflow:auto; min-width:0; }
  .card{
    background:var(--card);
    border:1px solid var(--line); border-radius:16px; padding:14px;
    box-shadow:var(--elev);
  }
  /* Make the right column stick while scrolling; don’t let it get cramped */
  .screener{ position: sticky; top:14px; height: calc(100vh - 28px); overflow: auto; min-width: 0; }

  /* Widen table + panes naturally */
  .table-wrap{ overflow:auto; max-height:76vh; border-radius:12px }
  table{ border-collapse:separate; border-spacing:0; width:100%; background:var(--card);
         border:1px solid var(--line); border-radius:14px; overflow:hidden; box-shadow: var(--elev); }
  /* ======= HEADINGS / CONTROLS (unchanged look) ======= */
  .h{ font-size:18px; margin:0 0 10px; font-weight:800; letter-spacing:.06em; text-transform:uppercase; }
  .note{ color:var(--muted); font-size:12px }
  .row{ display:flex; gap:10px; align-items:center; flex-wrap:wrap }
  input,select{
    background:#fff; border:1px solid var(--line); color:var(--ink);
    border-radius:12px; padding:9px 12px; outline:none; transition:box-shadow .2s, border-color .2s;
  }
  input:focus, select:focus{ box-shadow: var(--focus); border-color: #dccaa9; }
  .btn{
    border:1px solid #d9d1c3; background:linear-gradient(180deg,#fff, #f7f3eb);
    color:#1b1b1d; padding:9px 12px; border-radius:12px; cursor:pointer; font-weight:700;
    transition: transform .06s ease, box-shadow .2s, border-color .2s, background .2s;
    box-shadow: 0 1px 0 rgba(0,0,0,.02), 0 6px 16px rgba(0,0,0,.06);
  }
  .btn:hover{ transform: translateY(-1px); background:linear-gradient(180deg,#fff,#f3eee5) }
  .btn:active{ transform: translateY(0) }
  .btn.secondary{ background:#fff; border:1px solid var(--line); color:#2a2a2c; }
  .btn:focus{ box-shadow: var(--focus) }
  .chips{ display:flex; gap:10px; flex-wrap:wrap }
  .chip{ display:flex; align-items:center; gap:10px; background:#fff; border:1px solid var(--line);
         padding:8px 10px; border-radius:999px; font-size:13px; box-shadow: inset 0 -1px 0 rgba(0,0,0,.02); }
  .seg{ display:flex; border:1px solid var(--line); border-radius:10px; overflow:hidden; background:#faf7f0 }
  .seg button{ all:unset; padding:6px 10px; cursor:pointer; font-weight:800; color:#2c2c2f; }
  .seg button+button{ border-left:1px solid var(--line) }
  .seg button.on{ background:linear-gradient(180deg,#ffe9be,#f8dfaa); color:#2a2110; }
  .tag{ display:inline-flex; align-items:center; gap:8px; background:#fff; border:1px solid var(--line);
        color:#1f1f21; padding:8px 12px; border-radius:999px; box-shadow: inset 0 -1px 0 rgba(0,0,0,.02); }
  .pill{ display:inline-block; min-width:64px; padding:5px 12px; border-radius:999px; border:1px solid var(--line);
         background:#ffffff; transition: box-shadow .2s, border-color .2s; }
  .pill:hover{ border-color:#d8c8a9; box-shadow:0 0 0 3px rgba(197,165,114,.18) }

  /* ======= TABLE (unchanged look) ======= */
  thead th{
    position:sticky; top:0; background:#faf9f6; font-weight:800; font-size:13px; color:#1b1b1d;
    text-align:left; border-bottom:1px solid var(--line); padding:11px 10px; letter-spacing:.02em;
  }
  tbody td{ border-bottom:1px solid var(--line); padding:11px 10px; font-size:14px }
  tbody tr:nth-child(even){ background:var(--band) }
  tbody tr:hover td{ background:#f7f3e9 }
  .t-right{text-align:right} .pos{color:#1f7d4d} .neg{color:#a33a2a} .muted{color:var(--muted)}

  /* ======= Typeahead ======= */
  .ta-wrap{ position:relative; display:inline-block }
  .ta-list{ position:absolute; top:100%; left:0; right:0; z-index:50; background:#fff; border:1px solid var(--line);
            border-radius:12px; margin-top:6px; max-height:280px; overflow:auto; box-shadow:var(--elev); display:none; }
  .ta-item{ padding:10px 12px; display:flex; gap:10px; align-items:center; cursor:pointer }
  .ta-item b{ font-weight:800; min-width:70px; letter-spacing:.02em }
  .ta-item small{ color:var(--muted) }
  .ta-item:hover{ background:#f7f3e9 }

  /* ======= Mirrored horizontal scrollbar (still works) ======= */
  #finHScroll{
    position: fixed; bottom: 10px; height: 16px; overflow-x: auto; overflow-y: hidden; z-index: 9999;
    background: rgba(255,255,255,.75); border:1px solid var(--line); border-radius: 10px; display:none;
    box-shadow: var(--elev); backdrop-filter: blur(6px);
  }
  #finHScroll::-webkit-scrollbar{ height: 12px; }
  #finHScroll::-webkit-scrollbar-thumb{ background:#d4c8b3; border-radius:8px; }
  #finHScrollInner{ height:1px; }

  /* ======= Responsive tweaks ======= */
  /* Slightly larger screener on ultra-wide monitors */
  @media (min-width: 1800px){
    .grid{ grid-template-columns: minmax(0, 1fr) clamp(380px, 24vw, 640px); }
  }
  /* Comfortable screener width on medium desktops */
  @media (max-width: 1200px){
    .grid{ grid-template-columns: minmax(0, 1fr) clamp(320px, 30vw, 480px); }
  }
  /* Stack on small screens; remove sticky to avoid crowding */
  @media (max-width: 980px){
    .grid{ grid-template-columns: 1fr; }
    .screener{ position: static; height:auto; }
  }

  .block{ border:1px solid var(--line); border-radius:14px; padding:10px; background:#fff; box-shadow: var(--elev); margin-bottom:8px }
.block-head{ display:flex; align-items:center; gap:8px; margin-bottom:6px }
.hint{ display:inline-flex; width:18px; height:18px; border-radius:50%; align-items:center; justify-content:center;
       font-size:12px; border:1px solid var(--line); color:#6b5f46; background:#faf7f0; cursor:help }
.adv.hide{ display:none }

.sector-block{ border:1px solid var(--line); border-radius:14px; padding:10px; background:#fff; box-shadow: var(--elev); margin-bottom:8px }
.sector-chips{ display:flex; flex-wrap:wrap; gap:8px; max-height:220px; overflow:auto; padding:4px }
.sector-chip{
  display:inline-flex; align-items:center; gap:8px; padding:8px 12px; border-radius:999px;
  border:1px solid var(--line); background:#ffffff; cursor:pointer; user-select:none;
  box-shadow: inset 0 -1px 0 rgba(0,0,0,.02);
}
.sector-chip:hover{ border-color:#d8c8a9; box-shadow:0 0 0 3px rgba(197,165,114,.18) }
.sector-chip.on{ background:linear-gradient(180deg,#ffe9be,#f8dfaa); color:#2a2110; border-color:#d3b57b }
</style>
    """
    source_label = meta_bs.get("source", "SEC")
    html = f"""
    <div id="fin-root">
      <div class="meta">Source: {source_label} • CIK {meta_bs.get('cik', '-')} • Years: {', '.join(meta_bs['years'])}</div>
      <div class="fin-wrap">
        {render_table("Balance Sheet", bs_rows, meta_bs['years'])}
        {render_table("Income Statement", is_rows, meta_is['years'])}
        {render_table("Cash Flow Statement", cf_rows, meta_cf['years'])}
      </div>
    </div>
    """
    return HTMLResponse(css + html, headers=NO_CACHE_HEADERS)


@app.get("/financials/{ticker}/structured.xlsx")
def structured_xlsx(ticker: str):
    ticker = _normalize_ticker(ticker)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except Exception:
        raise HTTPException(500, "XLSX export requires 'openpyxl'. Install with: pip install openpyxl")

    bs_rows, meta_bs = _build_matrix_10y_any(ticker, BALANCE_SHEET_LAYOUT)
    is_rows, meta_is = _build_matrix_10y_any(ticker, INCOME_STATEMENT_LAYOUT)
    cf_rows, meta_cf = _build_matrix_10y_any(ticker, CASH_FLOW_LAYOUT)

    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(style="thin", color="D9D9D9")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    head_fill = PatternFill("solid", fgColor="F6F7F9")
    band_fill = PatternFill("solid", fgColor="F2F4F7")
    total_fill = PatternFill("solid", fgColor="FBFBFC")

    def write_sheet(title: str, rows: list, years: list[str]):
        ws = wb.create_sheet(title=title)
        ws.cell(row=1, column=1, value="")
        for c, y in enumerate(years, start=2):
            cell = ws.cell(row=1, column=c, value=y)
            cell.font = Font(bold=True)
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="right")
            cell.border = border
        r = 2
        for row in rows:
            if row.get("banner") == "header":
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1 + len(years))
                cell = ws.cell(row=r, column=1, value=row["label"])
                cell.font = Font(bold=True)
                cell.fill = band_fill
                cell.border = border
                r += 1
                continue
            if row.get("banner") == "subheader":
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1 + len(years))
                cell = ws.cell(row=r, column=1, value=row["label"])
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="FAFAFA")
                cell.border = border
                r += 1
                continue
            lc = ws.cell(row=r, column=1, value=row["label"])
            lc.alignment = Alignment(horizontal="left", indent=row.get("indent", 0))
            lc.border = border
            if row.get("is_total"):
                lc.font = Font(bold=True)
                lc.fill = total_fill
            for c, v in enumerate(row["values"], start=2):
                cell = ws.cell(row=r, column=c)
                if v is None:
                    cell.value = "-"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = float(v)
                    cell.number_format = '#,##0;(#,##0);"-"'
                    cell.alignment = Alignment(horizontal="right")
                cell.border = border
                if row.get("is_total"):
                    cell.font = Font(bold=True)
                    cell.fill = total_fill
            r += 1
        ws.column_dimensions["A"].width = 42
        for i in range(2, 2 + len(years)):
            ws.column_dimensions[chr(64 + i)].width = 14
        ws.freeze_panes = "B2"

    write_sheet("Balance Sheet", bs_rows, meta_bs["years"])
    write_sheet("Income Statement", is_rows, meta_is["years"])
    write_sheet("Cash Flow", cf_rows, meta_cf["years"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {
        "Content-Disposition": f'attachment; filename="{meta_bs["ticker"]}_financials_10y.xlsx"',
        **NO_CACHE_HEADERS,
    }
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/financials/{ticker}/structured.csv")
def structured_csv(ticker: str):
    ticker = _normalize_ticker(ticker)
    bs_rows, meta_bs = _build_matrix_10y_any(ticker, BALANCE_SHEET_LAYOUT)
    is_rows, meta_is = _build_matrix_10y_any(ticker, INCOME_STATEMENT_LAYOUT)
    cf_rows, meta_cf = _build_matrix_10y_any(ticker, CASH_FLOW_LAYOUT)

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["statement", "section", "label", "year", "value"])
    writer.writeheader()

    def write_rows(name, rows, years):
        section = ""
        for r in rows:
            if r.get("banner") in ("header", "subheader"):
                section = r["label"]
                continue
            for y, v in zip(years, r["values"]):
                writer.writerow({"statement": name, "section": section, "label": r["label"], "year": y, "value": v})

    write_rows("Balance Sheet", bs_rows, meta_bs["years"])
    write_rows("Income Statement", is_rows, meta_is["years"])
    write_rows("Cash Flow Statement", cf_rows, meta_cf["years"])

    buf.seek(0)
    fname = f'{meta_bs["ticker"]}_financials_10y.csv'
    headers = {"Content-Disposition": f'attachment; filename="{fname}"', **NO_CACHE_HEADERS}
    return StreamingResponse(iter([buf.read()]), media_type="text/csv", headers=headers)


# -------- Combined UI (left: financials, right: screener) --------
@app.get("/", response_class=HTMLResponse)
def home():
    return HTMLResponse(
        r"""<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <title>Global Screener + Financials</title>
  <style>
    /* ===== Quiet-lux system-light theme ===== */
    :root{
      --bg:#f6f5f2;          /* ivory canvas */
      --card:#ffffff;
      --ink:#111111;         /* deep charcoal text */
      --muted:#6e6e73;       /* subtle gray */
      --line:#e6e2d9;        /* hairline border */
      --band:#fbfaf7;        /* zebra band */
      --elev:0 12px 32px rgba(0,0,0,.06);
      --accent:#c5a572;      /* champagne gold */
      --accent-ink:#2a2110;  /* dark for contrast when needed */
      --focus:0 0 0 3px rgba(197,165,114,.25); /* gold focus ring */
    }
    html,body{
      background:var(--bg); color:var(--ink);
      font-family:ui-sans-serif,-apple-system,BlinkMacSystemFont,"SF Pro Text",Inter,Segoe UI,Roboto,Arial,sans-serif;
      font-size:15px; letter-spacing:.01em; margin:0; height:100%;
    }

    /* Layout */
    .grid{
      display:grid; grid-template-columns:minmax(0,1fr) 420px; gap:16px;
      max-width:1200px; margin:14px auto; padding:0 14px; align-items:start;
    }
    .card{
      background:var(--card);
      border:1px solid var(--line); border-radius:16px; padding:14px;
      box-shadow:var(--elev);
    }

    /* Headings */
    .h{
      font-size:18px; margin:0 0 10px; font-weight:800;
      letter-spacing:.06em; text-transform:uppercase;
    }
    .note{ color:var(--muted); font-size:12px }
    /* Controls */
    .row{ display:flex; gap:10px; align-items:center; flex-wrap:wrap }
    input,select{
      background:#fff; border:1px solid var(--line); color:var(--ink);
      border-radius:12px; padding:9px 12px; outline:none; transition:box-shadow .2s, border-color .2s;
    }
    input:focus, select:focus{ box-shadow: var(--focus); border-color: #dccaa9; }
    .btn{
      border:1px solid #d9d1c3; background:
        linear-gradient(180deg,#fff, #f7f3eb);
      color:#1b1b1d; padding:9px 12px; border-radius:12px; cursor:pointer; font-weight:700;
      transition: transform .06s ease, box-shadow .2s, border-color .2s, background .2s;
      box-shadow: 0 1px 0 rgba(0,0,0,.02), 0 6px 16px rgba(0,0,0,.06);
    }
    .btn:hover{ transform: translateY(-1px); background:linear-gradient(180deg,#fff,#f3eee5) }
    .btn:active{ transform: translateY(0) }
    .btn.secondary{
      background:#fff; border:1px solid var(--line); color:#2a2a2c;
    }
    .btn:focus{ box-shadow: var(--focus) }
    /* Chips / segmented controls */
    .chips{ display:flex; gap:10px; flex-wrap:wrap }
    .chip{
      display:flex; align-items:center; gap:10px;
      background:#fff; border:1px solid var(--line); padding:8px 10px; border-radius:999px; font-size:13px;
      box-shadow: inset 0 -1px 0 rgba(0,0,0,.02);
    }
    .seg{ display:flex; border:1px solid var(--line); border-radius:10px; overflow:hidden; background:#faf7f0 }
    .seg button{
      all:unset; padding:6px 10px; cursor:pointer; font-weight:800; color:#2c2c2f;
    }
    .seg button+button{ border-left:1px solid var(--line) }
    .seg button.on{
      background:linear-gradient(180deg,#ffe9be,#f8dfaa);
      color:#2a2110; /* accent ink */
    }
    /* Tags & pills */
    .tag{
      display:inline-flex; align-items:center; gap:8px;
      background:#fff; border:1px solid var(--line); color:#1f1f21;
      padding:8px 12px; border-radius:999px; box-shadow: inset 0 -1px 0 rgba(0,0,0,.02);
    }
    .pill{
      display:inline-block; min-width:64px; padding:5px 12px; border-radius:999px;
      border:1px solid var(--line); background:#ffffff;
      transition: box-shadow .2s, border-color .2s;
    }
    .pill:hover{ border-color:#d8c8a9; box-shadow:0 0 0 3px rgba(197,165,114,.18) }
    /* Table */
    .table-wrap{ overflow:auto; max-height:76vh; border-radius:12px }
    table{
      border-collapse:separate; border-spacing:0; width:100%; background:var(--card);
      border:1px solid var(--line); border-radius:14px; overflow:hidden; box-shadow: var(--elev);
    }
    thead th{
      position:sticky; top:0; background:#faf9f6; font-weight:800; font-size:13px; color:#1b1b1d;
      text-align:left; border-bottom:1px solid var(--line); padding:11px 10px; letter-spacing:.02em;
    }
    tbody td{ border-bottom:1px solid var(--line); padding:11px 10px; font-size:14px }
    tbody tr:nth-child(even){ background:var(--band) }
    tbody tr:hover td{ background:#f7f3e9 } /* luxe hover */
    .t-right{text-align:right}
    .pos{color:#1f7d4d} .neg{color:#a33a2a}
    .muted{color:var(--muted)}
    /* Screener column stickiness and panel */
    .screener { position: sticky; top:14px; height: calc(100vh - 28px); overflow: auto; }
    /* Typeahead */
    .ta-wrap{ position:relative; display:inline-block }
    .ta-list{
      position:absolute; top:100%; left:0; right:0; z-index:50; background:#fff; border:1px solid var(--line);
      border-radius:12px; margin-top:6px; max-height:280px; overflow:auto; box-shadow:var(--elev); display:none;
    }
    .ta-item{ padding:10px 12px; display:flex; gap:10px; align-items:center; cursor:pointer }
    .ta-item b{ font-weight:800; min-width:70px; letter-spacing:.02em }
    .ta-item small{ color:var(--muted) }
    .ta-item:hover{ background:#f7f3e9 }
    /* Mirrored scrollbar (unchanged behavior, themed) */
    #finHScroll{ position: fixed; bottom: 10px; height: 16px; overflow-x: auto; overflow-y: hidden; z-index: 9999;
                 background: rgba(255,255,255,.75); border:1px solid var(--line); border-radius: 10px; display:none;
                 box-shadow: var(--elev); backdrop-filter: blur(6px); }
    #finHScroll::-webkit-scrollbar{ height: 12px; }
    #finHScroll::-webkit-scrollbar-thumb{ background:#d4c8b3; border-radius:8px; }
    #finHScrollInner{ height:1px; }
    /* TradingView-style compact overrides */
.toolbar select { margin-right: 8px; }
.chip { border-radius: 10px; padding: 6px 8px; }
.seg button { padding: 6px 8px; }
  </style>
</head>
<body>
  <div class="grid">
    <!-- LEFT: FINANCIALS -->
    <section class="card" id="finCard">
      <div class="fin-head row" style="margin-bottom:8px">
        <h2 class="h">Financials</h2>
        <div class="ta-wrap">
          <input id="tickerInput" placeholder="Type ticker or name…" style="width:260px" aria-label="Ticker search">
          <div id="taList" class="ta-list" role="listbox" aria-label="Suggestions"></div>
        </div>
        <select id="periodSelect" title="Period">
  <option value="annual" selected>Annual</option>
  <option value="quarterly">Quarterly</option>
</select>

<select id="unitsSelect" title="Units">
  <option value="billions">Billions</option>
  <option value="millions" selected>Millions</option>
  <option value="thousands">Thousands</option>
</select>
        <span class="note" id="finStatus"></span>
        <span style="flex:1"></span>
        <input id="finSearch" placeholder="Search rows…" style="width:220px" aria-label="Filter lines">
        <label class="note" title="Keep current ticker shown on hover"><input type="checkbox" id="pinHover"> Pin ticker</label>
      </div>
      <div id="finPane" class="table-wrap" style="max-height: calc(100vh - 160px);"></div>
    </section>
    <!-- RIGHT: SCREENER -->
    <aside class="card screener">
      <h2 class="h">Screener</h2>
      <div class="row" style="margin-bottom:6px">
        <b style="letter-spacing:.02em">Universe</b>
        <button class="btn" onclick="bootstrapUS()">US</button>
        <button class="btn secondary" onclick="bootstrapGlobal()">Global</button>
        <button class="btn secondary" onclick="addCustom()">Add custom</button>
        <span class="note" id="ucount"></span>
      </div>

      
      

      <div class="row" style="margin-bottom:6px">
        <div class="chips" id="chips">
          <div class="chip" data-field="market_cap">
            <span>Market Cap</span>
            <div class="seg">
              <button data-op=">" class="on">&gt;</button>
              <button data-op=">=">&ge;</button>
              <button data-op="<">&lt;</button>
              <button data-op="<=">&le;</button>
            </div>
            <input class="val" value="1e11" style="width:110px"/>
          </div>
          <div class="chip" data-field="pe">
            <span>P/E</span>
            <div class="seg">
              <button data-op="<" class="on">&lt;</button>
              <button data-op="<=">&le;</button>
              <button data-op=">">&gt;</button>
              <button data-op=">=">&ge;</button>
            </div>
            <input class="val" value="40" style="width:88px"/>
          </div>
        </div>
      </div>

      <div class="row" style="margin-bottom:6px">
  <span class="note">Rows</span><input id="limUI" value="100" style="width:80px"/>
  <button class="btn secondary" onclick="prevPage()">Prev</button>
  <button class="btn secondary" onclick="nextPage()">Next</button>
  <span class="note" id="pageInfo"></span>
</div>

      <div class="sector-block">
  <div class="block-head">
    <b>Sector</b>
    <span class="note">Click a chip to load top 200 by market cap. Shift/ctrl-click to select multiple for filtering.</span>
  </div>
  <div id="sectorChips" class="sector-chips"></div>
</div>

      <div class="table-wrap">
        <table id="tbl">
          <thead>
            <tr>
              <th data-col="ticker">Ticker</th>
              <th data-col="name">Name</th>
              <th data-col="sector">Sector</th>
              <th data-col="industry">Industry</th>
              <th class="t-right" data-col="price">Price</th>
              <th class="t-right" data-col="market_cap">Market cap</th>
              <th class="t-right" data-col="pe">P/E</th>
              <th class="t-right" data-col="ps">P/S</th>
              <th class="t-right" data-col="pb">P/B</th>
              <th class="t-right" data-col="rev_cagr_3y">Rev CAGR 3Y</th>
              <th class="t-right" data-col="net_margin">Net margin</th>
              <th class="t-right" data-col="debt_to_equity">D/E</th>
              <th class="t-right" data-col="updated_at">Updated</th>
            </tr>
          </thead>
          <tbody id="tbody"></tbody>
        </table>
      </div>
    </aside>
  </div>

  <div id="finHScroll"><div id="finHScrollInner"></div></div>

<script>
/* ===== formatting helpers ===== */
const nf2 = new Intl.NumberFormat('en-US', {maximumFractionDigits: 2});
const compact = new Intl.NumberFormat('en-US',{notation:'compact', maximumFractionDigits:2});
const fmt = {
  money: (n) => n==null? '-' : '$' + compact.format(n),
  num:   (n) => n==null? '-' : nf2.format(n),
  ratio: (n) => n==null? '-' : nf2.format(n),
  pct:   (n) => n==null? '-' : `<span class="${n>=0?'pos':'neg'}">${nf2.format(n*100)}%</span>`,
  date:  (s) => s? `<span class="muted">${s.replace('T',' ').replace('Z','')}</span>` : '-'
};
function debounce(fn, ms=300){ let t; return (...a)=>{ clearTimeout(t); t=setTimeout(()=>fn(...a), ms) } }

/* ===== Advanced toggle ===== */
function toggleAdvanced(){
  const el = document.getElementById('advRefresh');
  el?.classList.toggle('hide');
}

/* ===== Sector chips ===== */
function makeSectorChip(name){
  const div = document.createElement('div');
  div.className = 'sector-chip';
  div.textContent = name;
  div.dataset.value = name;
  div.addEventListener('click', async ()=>{
    div.classList.toggle('on');       // toggle selection (filters)
    run();                            // apply filters
    try{                              // one-click preview: top 200 in sector
      const r = await fetch(`/sector/top?sector=${encodeURIComponent(name)}&n=200`);
      const rows = await r.json();
      render(rows);
      document.getElementById('pageInfo').innerText = `Top in ${name} • ${rows.length}`;
    }catch(e){}
  });
  return div;
}
async function buildSectorChips(){
  try{
    const r = await fetch('/meta', { cache:'no-store' });
    const js = await r.json();
    const wrap = document.getElementById('sectorChips');
    if(!wrap) return;
    wrap.innerHTML = '';
    (js.sectors||[]).forEach(s => wrap.appendChild(makeSectorChip(s)));
  }catch(e){}
}
function selectedSectors(){
  return [...document.querySelectorAll('.sector-chip.on')].map(el => el.dataset.value);
}

/* ===== Screener helpers ===== */
function opForChip(ch){ return ch.querySelector('.seg .on')?.dataset.op || '>'; }
function valueForChip(ch){
  const raw = ch.querySelector('.val')?.value?.trim() ?? '';
  if(raw==='') return null;
  const num = Number(raw);
  return isNaN(num) ? raw : num;
}
function buildFilters(){
  const filters=[];
  document.querySelectorAll('#chips .chip').forEach(ch=>{
    const field = ch.dataset.field, op = opForChip(ch);
    const val = valueForChip(ch);
    if(val!==null) filters.push({field, op, value: val});
  });
  const sectors = selectedSectors();
  if(sectors.length) filters.push({field:'sector', op:'in', value: sectors});
  return filters;
}
async function refreshCount(){
  try{
    const r=await fetch('/universe/count');
    const js=await r.json();
    document.getElementById('ucount').innerText=(js.count||0)+' symbols';
  }catch(e){}
}
async function bootstrapUS(){ const r=await fetch('/universe/bootstrap_us',{method:'POST'}); const js=await r.json(); await refreshCount(); alert('US bootstrap added: '+(js.added||0)); }
async function bootstrapGlobal(){ const r=await fetch('/universe/bootstrap_global',{method:'POST'}); const js=await r.json(); await refreshCount(); alert('Global added: '+(js.total_added||0)); }
async function addCustom(){
  const ex=prompt('Enter Yahoo symbols separated by commas (e.g., RY.TO,7203.T,0700.HK):'); if(!ex) return;
  const arr=ex.split(',').map(s=>s.trim()).filter(Boolean).map(t=>({ticker:t, active:true}));
  await fetch('/universe/add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(arr)});
  await refreshCount();
}
async function ingestSome(){
  const t=(document.getElementById('tickers').value||'').split(',').map(s=>s.trim()).filter(Boolean);
  if(!t.length) return; const url='/ingest?'+t.map(x=>'tickers='+encodeURIComponent(x)).join('&');
  await fetch(url,{method:'POST'}); run();
}
async function ingestAll(){
  const ttl=+document.getElementById('ttl').value||24;
  const lim=+document.getElementById('lim').value||1500;
  const par=+document.getElementById('par').value||8;
  document.getElementById('ingmsg').innerText='Running...';
  const r=await fetch(`/ingest_all?stale_hours=${ttl}&limit=${lim}&parallel=${par}`,{method:'POST'});
  const js=await r.json();
  document.getElementById('ingmsg').innerText=`selected ${js.selected}, ok ${js.ok}, errors ${js.errors}`;
  run();
}

/* ===== Screener run/render ===== */
let currentSort = {field:'market_cap', dir:'desc'};
let page = 0;

async function run(){
  const limit = +document.getElementById('limUI').value||100;
  const offset = page*limit;
  const body = {filters: buildFilters(), sort:[currentSort], limit, offset};
  const url  = `/screen?auto_fill=1&ttl_hours=24&parallel=8`;
  const res = await fetch(url,{method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(body)});
  const rows = await res.json();
  render(rows);
  document.getElementById('pageInfo').innerText = `page ${page+1} • showing ${rows.length}`;
}
function nextPage(){ page++; run(); }
function prevPage(){ page=Math.max(0,page-1); run(); }

function render(rows){
  const {field, dir} = currentSort;
  rows.sort((a,b)=>{
    const av=a[field], bv=b[field];
    if(av==null && bv==null) return 0;
    if(av==null) return 1;
    if(bv==null) return -1;
    if(typeof av==='string' || typeof bv==='string'){
      return dir==='asc' ? String(av).localeCompare(String(bv)) : String(bv).localeCompare(String(av));
    }
    return dir==='asc' ? (av-bv) : (bv-av);
  });
  const tb = document.getElementById('tbody');
  tb.innerHTML = rows.map(r => `
    <tr class="row-stock" data-ticker="${r.ticker||''}" title="${r.name||''}">
      <td data-col="ticker"><a href="https://finance.yahoo.com/quote/${r.ticker}" target="_blank" class="pill" style="text-decoration:none;color:inherit">${r.ticker||'-'}</a></td>
      <td data-col="name">${r.name||'-'}</td>
      <td data-col="sector" class="muted">${r.sector||'-'}</td>
      <td data-col="industry" class="muted">${r.industry||'-'}</td>
      <td data-col="price" class="t-right">${fmt.num(r.price)}</td>
      <td data-col="market_cap" class="t-right" title="${r.market_cap??''}">${fmt.money(r.market_cap)}</td>
      <td data-col="pe" class="t-right">${fmt.ratio(r.pe)}</td>
      <td data-col="ps" class="t-right">${fmt.ratio(r.ps)}</td>
      <td data-col="pb" class="t-right">${fmt.ratio(r.pb)}</td>
      <td data-col="rev_cagr_3y" class="t-right">${fmt.pct(r.rev_cagr_3y)}</td>
      <td data-col="net_margin" class="t-right">${fmt.pct(r.net_margin)}</td>
      <td data-col="debt_to_equity" class="t-right">${fmt.ratio(r.debt_to_equity)}</td>
      <td data-col="updated_at" class="t-right">${fmt.date(r.updated_at)}</td>
    </tr>
  `).join('');

  // hover-to-preview financials
  let hoverTimer = null;
  tb.querySelectorAll('tr.row-stock').forEach(tr=>{
    tr.addEventListener('mouseenter', ()=>{
      if(document.getElementById('pinHover')?.checked) return;
      const t = tr.dataset.ticker;
      clearTimeout(hoverTimer);
      hoverTimer = setTimeout(()=> loadFinancials(t, {fromHover:true}), 200);
    });
  });
  tb.addEventListener('mouseleave', ()=> clearTimeout(hoverTimer));
}

/* ===== Type-ahead ===== */
const ta = { wrap:null, list:null, input:null, items:[], index:-1, open:false };
function taRender(){
  const el = ta.list; el.innerHTML = '';
  if(!ta.items.length){ el.style.display='none'; ta.open=false; return; }
  ta.items.forEach((it, i)=>{
    const row = document.createElement('div');
    row.className = 'ta-item' + (i===ta.index?' active':'');
    row.setAttribute('role','option');
    row.innerHTML = `<b>${it.symbol}</b><span>${it.name||''}</span><small>${it.exchange||''}${it.sector? ' • '+it.sector:''}</small>`;
    row.addEventListener('mousedown', (e)=>{ e.preventDefault(); taPick(i); });
    el.appendChild(row);
  });
  el.style.display='block'; ta.open=true;
}
function taPick(i){
  const it = ta.items[i]; if(!it) return;
  ta.input.value = it.symbol;
  ta.list.style.display='none'; ta.open=false;
  loadFinancials(it.symbol);
}
const fetchSuggest = debounce(async (q)=>{
  if(!q || q.length<1){ ta.items=[]; taRender(); return; }
  try{
    const r = await fetch(`/suggest?q=${encodeURIComponent(q)}&limit=8`, { cache:'no-store' });
    const js = await r.json();
    ta.items = (js.matches||[]); ta.index = -1; taRender();
  }catch(_){ ta.items=[]; taRender(); }
}, 180);

/* ===== Financials pane ===== */
let currentTicker = null, finAbort = null;
const finCache = new Map(); const FIN_CACHE_MAX = 24; const FIN_CACHE_TTL_MS = 30*1000;

function setFinHTML(html){
  const pane = document.getElementById('finPane');
  pane.innerHTML = html;
  applyFinSearch();
  ensureFinScrollbar();
}
async function loadFinancials(ticker, { fromHover=false } = {}){
  if(!ticker) return;
  ticker = ticker.trim().toUpperCase();
  if(fromHover && ticker === currentTicker) return;

  currentTicker = ticker;
  document.getElementById('tickerInput').value = ticker;
  document.getElementById('finStatus').innerText = 'Loading ' + ticker + '…';

  const cached = finCache.get(ticker);
  if (cached && (Date.now() - cached.ts) < FIN_CACHE_TTL_MS) {
    setFinHTML(cached.html);
    document.getElementById('finStatus').innerText = `Showing ${ticker}`;
    return;
  } else {
    finCache.delete(ticker);
  }

  if(finAbort) finAbort.abort();
  finAbort = new AbortController();
  try{
    const freq  = (document.getElementById('periodSelect')?.value || 'Annual').toLowerCase();
const units = (document.getElementById('unitsSelect')?.value  || 'Millions').toLowerCase();
const url = `/financials/${encodeURIComponent(ticker)}/structured.html?freq=${freq}&units=${units}&ts=${Date.now()}`;
    const res = await fetch(url, { signal: finAbort.signal, cache: 'no-store' });
    if(!res.ok) throw new Error(`HTTP ${res.status}`);
    const html = await res.text();
    finCache.set(ticker, { html, ts: Date.now() });
    if(finCache.size > FIN_CACHE_MAX){ const firstKey = finCache.keys().next().value; finCache.delete(firstKey); }
    setFinHTML(html);
    document.getElementById('finStatus').innerText = `Showing ${ticker}`;
  } catch(e){
    if(e.name==='AbortError') return;
    document.getElementById('finPane').innerHTML = `<div class="note">Failed to load ${ticker}: ${e.message}</div>`;
    document.getElementById('finStatus').innerText = '';
  } finally { finAbort = null; }
}
document.getElementById('finSearch').addEventListener('input', debounce(applyFinSearch, 200));
function applyFinSearch(){
  const term = (document.getElementById('finSearch').value||'').trim().toLowerCase();
  const root = document.getElementById('finPane'); if(!root) return;
  root.querySelectorAll('mark').forEach(m=>{ m.replaceWith(document.createTextNode(m.textContent)); });
  const allBodyRows = [...root.querySelectorAll('table.fin-table tbody tr')];
  allBodyRows.forEach(tr=>{
    tr.classList.remove('hide-row');
    const labelCell = tr.querySelector('td.label'); if(!labelCell) return;
    const txt = labelCell.textContent.toLowerCase();
    if(term && !txt.includes(term)){ tr.classList.add('hide-row'); }
    else if(term){
      const rx = new RegExp('('+term.replace(/[.*+?^${}()|[\]\\]/g,'\\$&')+')','ig');
      labelCell.innerHTML = labelCell.textContent.replace(rx,'<mark>$1</mark>');
    }
  });
  root.querySelectorAll('table.fin-table').forEach(tbl=>{
    const trs = [...tbl.querySelector('tbody').children];
    for(let i=0;i<trs.length;i++){
      const tr = trs[i];
      if(tr.classList.contains('section') || tr.classList.contains('subsection')){
        let j=i+1, any=false;
        while(j<trs.length && !trs[j].classList.contains('section') && !trs[j].classList.contains('subsection')){
          if(!trs[j].classList.contains('hide-row')) { any = true; break; }
          j++;
        }
        tr.style.display = any ? '' : 'none';
      }
    }
  });
}

/* ===== mirrored horizontal scrollbar ===== */
let finScrollSync = { wrap:null, onWrap:null, onBar:null, syncing:false };
function getFinWrap(){ return document.querySelector('#finPane .fin-wrap'); }
function updateFinScrollbarGeometry(){
  const card = document.getElementById('finCard'), bar = document.getElementById('finHScroll'), inner = document.getElementById('finHScrollInner');
  const wrap = getFinWrap(); if(!card || !bar || !inner || !wrap){ bar.style.display='none'; return; }
  const rect = card.getBoundingClientRect(), pad=14;
  bar.style.left  = (rect.left + pad) + 'px';
  bar.style.width = Math.max(200, rect.width - pad*2) + 'px';
  const needs = wrap.scrollWidth > wrap.clientWidth + 1;
  bar.style.display = needs ? 'block' : 'none';
  inner.style.width = wrap.scrollWidth + 'px';
  if(!finScrollSync.syncing){ bar.scrollLeft = wrap.scrollLeft; }
}
function ensureFinScrollbar(){
  const wrap = getFinWrap(), bar = document.getElementById('finHScroll'), inner = document.getElementById('finHScrollInner');
  if(!wrap || !bar || !inner) return;
  if(finScrollSync.wrap && finScrollSync.onWrap){
    finScrollSync.wrap.removeEventListener('scroll', finScrollSync.onWrap);
    bar.removeEventListener('scroll', finScrollSync.onBar);
  }
  finScrollSync.wrap = wrap;
  finScrollSync.onWrap = ()=>{ if(finScrollSync.syncing) return; finScrollSync.syncing=true; bar.scrollLeft = wrap.scrollLeft; finScrollSync.syncing=false; };
  finScrollSync.onBar  = ()=>{ if(finScrollSync.syncing) return; finScrollSync.syncing=true; wrap.scrollLeft = bar.scrollLeft; finScrollSync.syncing=false; };
  wrap.addEventListener('scroll', finScrollSync.onWrap, {passive:true});
  bar.addEventListener('scroll',  finScrollSync.onBar,  {passive:true});
  updateFinScrollbarGeometry();
}
window.addEventListener('resize', debounce(updateFinScrollbarGeometry, 100));
window.addEventListener('scroll', debounce(updateFinScrollbarGeometry, 50));

/* ===== Init wiring ===== */
document.addEventListener('click', (e)=>{
  const th = e.target.closest('th[data-col]');
  if(th){
    const col = th.dataset.col;
    currentSort = currentSort.field===col ? {field: col, dir:(currentSort.dir==='asc'?'desc':'asc')} : {field: col, dir:'asc'};
    run(); return;
  }
  const btn = e.target.closest('#chips .seg button');
  if(btn){
    const seg = btn.parentElement;
    seg.querySelectorAll('button').forEach(b=>b.classList.toggle('on', b===btn));
    run();
  }
});
document.querySelectorAll('#chips .val').forEach(inp=> inp.addEventListener('input', debounce(run, 250)));
document.getElementById('limUI')?.addEventListener('change', ()=>{ page=0; run(); });

/* Typeahead wiring */
document.addEventListener('DOMContentLoaded', ()=>{
  const input = document.getElementById('tickerInput');
  const list  = document.getElementById('taList');
  ta.input = input; ta.list = list; ta.wrap = input?.parentElement;
  if(ta.input && ta.list){
    ta.input.addEventListener('input', (e)=> fetchSuggest(e.target.value.trim()));
    ta.input.addEventListener('keydown', (e)=>{
      if(!ta.open && (e.key==='ArrowDown' || e.key==='ArrowUp')){ ta.open=true; taRender(); }
      if(!ta.open) { if(e.key==='Enter'){ loadFinancials(ta.input.value.trim()); } return; }
      if(e.key==='ArrowDown'){ e.preventDefault(); ta.index = (ta.index+1) % ta.items.length; taRender(); }
      else if(e.key==='ArrowUp'){ e.preventDefault(); ta.index = (ta.index-1+ta.items.length) % ta.items.length; taRender(); }
      else if(e.key==='Enter'){ if(ta.index>-1){ e.preventDefault(); taPick(ta.index); } else { loadFinancials(ta.input.value.trim()); } }
      else if(e.key==='Escape'){ ta.list.style.display='none'; ta.open=false; }
    });
    document.addEventListener('click', (e)=>{ if(!ta.wrap.contains(e.target)){ ta.list.style.display='none'; ta.open=false; } });
  }

  buildSectorChips();
  refreshCount();
    // Wire dropdowns: reload current ticker when user changes Period/Units
  const pSel = document.getElementById('periodSelect');
  const uSel = document.getElementById('unitsSelect');
  const _current = ()=> (document.getElementById('tickerInput').value || 'AAPL').trim();
  if (pSel) pSel.addEventListener('change', ()=> loadFinancials(_current()));
  if (uSel) uSel.addEventListener('change', ()=> loadFinancials(_current()));

  // First-run: if universe is empty, bootstrap all US tickers automatically
  try {
    const c = await (await fetch('/universe/count', {cache:'no-store'})).json();
    if ((c.count||0) === 0) {
      const res = await fetch('/universe/bootstrap_us', {method:'POST'});
      if (res.ok) {
        await refreshCount();   // update the counter
        run();                  // re-run screener to pull from the new universe
      }
    }
  } catch (_) {}
  loadFinancials('AAPL');
  run();
  requestAnimationFrame(()=> updateFinScrollbarGeometry());
});

/* Expose some functions used by buttons */
window.bootstrapUS = bootstrapUS;
window.bootstrapGlobal = bootstrapGlobal;
window.addCustom = addCustom;
window.ingestSome = ingestSome;
window.toggleAdvanced = toggleAdvanced;
window.ingestAll = ingestAll;
window.nextPage = nextPage;
window.prevPage = prevPage;
window.downloadXLSX = ()=>{ const t=(document.getElementById('tickerInput').value||'').trim(); if(!t) return; window.open(`/financials/${encodeURIComponent(t)}/structured.xlsx`, '_blank'); };
window.downloadCSV  = ()=>{ const t=(document.getElementById('tickerInput').value||'').trim(); if(!t) return; window.open(`/financials/${encodeURIComponent(t)}/structured.csv`, '_blank'); };
</script>
</body>
</html>
""",
        headers=NO_CACHE_HEADERS,
    )


# ---------- convenient runner ----------
if __name__ == "__main__":
    import uvicorn

    module_name = Path(__file__).stem
    uvicorn.run(f"{module_name}:app", host="127.0.0.1", port=8000, reload=True)